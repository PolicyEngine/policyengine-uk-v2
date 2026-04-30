"""Parse OBR EFO March 2026 detailed forecast tables into calibration targets.

Sources (local xlsx files in data/obr/):
- Receipts: efo-march-2026-detailed-forecast-tables-receipts.xlsx
- Expenditure: efo-march-2026-detailed-forecast-tables-expenditure.xlsx
- Economy: efo-march-2026-detailed-forecast-tables-economy.xlsx
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
OBR_DIR = REPO_ROOT / "data" / "obr"

RECEIPTS_FILE = OBR_DIR / "efo-march-2026-detailed-forecast-tables-receipts.xlsx"
EXPENDITURE_FILE = OBR_DIR / "efo-march-2026-detailed-forecast-tables-expenditure.xlsx"
ECONOMY_FILE = OBR_DIR / "efo-march-2026-detailed-forecast-tables-economy.xlsx"

# Sheet 3.8 (cash receipts): D=2024-25, E=2025-26, ..., J=2030-31
_RECEIPTS_COL_TO_YEAR = {
    "D": 2024,
    "E": 2025,
    "F": 2026,
    "G": 2027,
    "H": 2028,
    "I": 2029,
    "J": 2030,
}

# Sheet 4.9 (welfare): C=2024-25, D=2025-26, ..., I=2030-31
_WELFARE_COL_TO_YEAR = {
    "C": 2024,
    "D": 2025,
    "E": 2026,
    "F": 2027,
    "G": 2028,
    "H": 2029,
    "I": 2030,
}

# Sheet 4.1 (council tax): same layout as 4.9
_CT_COL_TO_YEAR = _WELFARE_COL_TO_YEAR

# Sheet 3.4 (IT + NICs detail): C=2024-25, ..., I=2030-31
_IT_COL_TO_YEAR = {
    "C": 2024,
    "D": 2025,
    "E": 2026,
    "F": 2027,
    "G": 2028,
    "H": 2029,
    "I": 2030,
}


def get_income_growth_indexes() -> dict[str, dict[int, float]]:
    """Return cumulative growth indexes relative to 2023 for each income type.

    Uses OBR sheet 3.5 (self-employment, dividend, property, savings growth
    rates) and sheet 1.6 (wages & salaries levels) to build indexes that can
    scale the HMRC SPI 2022-23 snapshot to other years.

    Returns e.g. {"employment_income": {2023: 1.0, 2024: 1.07, ...}, ...}
    """
    indexes: dict[str, dict[int, float]] = {}

    # ── Wages & salaries from sheet 1.6 (levels, £bn) ──
    if ECONOMY_FILE.exists():
        wb = openpyxl.load_workbook(ECONOMY_FILE, data_only=True)
        ws = wb["1.6"]
        wage_levels: dict[int, float] = {}
        for row in range(4, 200):
            b = ws.cell(row=row, column=2).value
            if b is None:
                continue
            year = _parse_fiscal_year(str(b))
            if year is not None and 2022 <= year <= 2030:
                val = ws.cell(row=row, column=14).value  # Col N = wages & salaries
                if val is not None and isinstance(val, (int, float)):
                    wage_levels[year] = float(val)
        wb.close()
        if 2022 in wage_levels:
            base = wage_levels[2022]
            indexes["employment_income"] = {y: v / base for y, v in wage_levels.items()}

    # ── Growth rates from sheet 3.5 ──
    # Cols: C=2023-24, D=2024-25, ..., J=2030-31
    _35_col_to_year = {3: 2023, 4: 2024, 5: 2025, 6: 2026, 7: 2027, 8: 2028, 9: 2029, 10: 2030}
    _35_rows = {
        "self_employment_income": 6,
        "dividend_income": 7,
        "property_income": 8,
        "savings_interest": 9,
    }
    if RECEIPTS_FILE.exists():
        wb = openpyxl.load_workbook(RECEIPTS_FILE, data_only=True)
        ws = wb["3.5"]
        for variable, data_row in _35_rows.items():
            # Build cumulative index from growth rates (% p.a.)
            # Base year is 2022 (FY 2022-23), so index[2022] = 1.0
            idx: dict[int, float] = {2022: 1.0}
            for col, year in sorted(_35_col_to_year.items()):
                rate = ws.cell(row=data_row, column=col).value
                if rate is not None and isinstance(rate, (int, float)):
                    prev_year = year - 1
                    idx[year] = idx.get(prev_year, 1.0) * (1 + rate / 100.0)
            indexes[variable] = idx
        wb.close()

    # State pension and private pension: use CPI as a proxy (triple lock ≈ max of CPI, AWE, 2.5%)
    # For calibration purposes CPI is a reasonable approximation
    if ECONOMY_FILE.exists():
        wb = openpyxl.load_workbook(ECONOMY_FILE, data_only=True)
        ws = wb["1.7"]
        # CPI growth is in a fiscal year row format too
        cpi_idx: dict[int, float] = {2022: 1.0}
        for row in range(4, 200):
            b = ws.cell(row=row, column=2).value
            if b is None:
                continue
            year = _parse_fiscal_year(str(b))
            if year is not None and 2023 <= year <= 2030:
                rate = ws.cell(row=row, column=4).value  # Col D = CPI
                if rate is not None and isinstance(rate, (int, float)):
                    cpi_idx[year] = cpi_idx.get(year - 1, 1.0) * (1 + rate / 100.0)
        wb.close()
        indexes["state_pension"] = cpi_idx
        indexes["private_pension_income"] = cpi_idx

    return indexes


def _find_row(ws, label: str, col: str = "B", max_row: int = 70) -> int | None:
    for row in range(1, max_row + 1):
        val = ws[f"{col}{row}"].value
        if val and str(val).strip().startswith(label):
            return row
    return None


def _read_row(ws, row: int, col_map: dict[str, int]) -> dict[int, float]:
    result = {}
    for col, year in col_map.items():
        val = ws[f"{col}{row}"].value
        if val is not None and isinstance(val, (int, float)):
            result[year] = float(val) * 1e9  # £bn → £
    return result


def _parse_receipts() -> list[dict]:
    """Parse sheet 3.8 for main tax receipts."""
    wb = openpyxl.load_workbook(RECEIPTS_FILE, data_only=True)
    ws = wb["3.8"]
    targets = []

    # Map: (label_prefix, target_name, variable in the survey data, entity, aggregation)
    # These are aggregate £ totals. For calibration we map them to survey-reported
    # income/benefit variables where possible.
    # Now that calibration runs after simulation, we can use simulated output
    # variables (income_tax, national_insurance, capital_gains_tax, etc.)
    receipt_rows = [
        (
            "Income tax (gross of tax credits)",
            "obr/income_tax_receipts",
            "income_tax",
            "person",
            "sum",
            "Simulated income tax",
        ),
        (
            "National insurance contributions",
            "obr/ni_receipts",
            "total_ni",
            "person",
            "sum",
            "Simulated employee + employer NI",
        ),
        (
            "Value added tax",
            "obr/vat_receipts",
            "vat",
            "household",
            "sum",
            "Simulated VAT",
        ),
        ("Fuel duties", "obr/fuel_duty_receipts", "fuel_duty", "household", "sum", ""),
        (
            "Capital gains tax",
            "obr/cgt_receipts",
            "capital_gains_tax",
            "household",
            "sum",
            "",
        ),
        (
            "Stamp duty land tax",
            "obr/sdlt_receipts",
            "stamp_duty",
            "household",
            "sum",
            "",
        ),
        (
            "Council tax",
            "obr/council_tax_receipts",
            "council_tax_annual",
            "household",
            "sum",
            "",
        ),
    ]

    for label, name, variable, entity, aggregation, _note in receipt_rows:
        if variable is None:
            continue  # Skip targets we can't map to survey data
        row = _find_row(ws, label)
        if row is None:
            continue
        values = _read_row(ws, row, _RECEIPTS_COL_TO_YEAR)
        for year, value in values.items():
            targets.append(
                {
                    "name": f"{name}/{year}",
                    "variable": variable,
                    "entity": entity,
                    "aggregation": aggregation,
                    "filter": None,
                    "value": value,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

    wb.close()
    return targets


def _parse_it_nics_detail() -> list[dict]:
    """Parse sheet 3.4 for income tax and NICs breakdown."""
    wb = openpyxl.load_workbook(RECEIPTS_FILE, data_only=True)
    ws = wb["3.4"]
    targets = []

    rows_to_parse = [
        (
            "Income tax (gross of tax credits)",
            "obr/income_tax",
            "income_tax",
            "person",
            "sum",
        ),
        (
            "Class 1 Employee NICs",
            "obr/ni_employee",
            "national_insurance",
            "person",
            "sum",
        ),
    ]

    for label, name, variable, entity, aggregation in rows_to_parse:
        row = _find_row(ws, label)
        if row is None:
            continue
        values = _read_row(ws, row, _IT_COL_TO_YEAR)
        for year, value in values.items():
            targets.append(
                {
                    "name": f"{name}/{year}",
                    "variable": variable,
                    "entity": entity,
                    "aggregation": aggregation,
                    "filter": None,
                    "value": value,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

    wb.close()
    return targets


def _parse_welfare() -> list[dict]:
    """Parse sheet 4.9 for benefit spending totals."""
    wb = openpyxl.load_workbook(EXPENDITURE_FILE, data_only=True)
    ws = wb["4.9"]
    targets = []

    # Map OBR row labels to simulated benefit variables.
    # Benefits are calculated at benunit level in the simulation.
    benefit_rows = [
        (
            "Housing benefit (not on JSA)",
            "obr/housing_benefit",
            "housing_benefit",
            "benunit",
        ),
        (
            "Disability living allowance and personal independence p",
            "obr/pip_dla",
            "pip_daily_living",
            "person",  # PIP/DLA are passthrough (input data), not simulated
        ),
        (
            "Attendance allowance",
            "obr/attendance_allowance",
            "attendance_allowance",
            "person",  # Passthrough
        ),
        ("Pension credit", "obr/pension_credit", "pension_credit", "benunit"),
        ("Carer's allowance", "obr/carers_allowance", "carers_allowance", "benunit"),
        ("Child benefit", "obr/child_benefit", "child_benefit", "benunit"),
        ("State pension", "obr/state_pension", "state_pension", "benunit"),
    ]

    # UC appears twice in 4.9 — inside and outside the welfare cap. Sum them
    # into a single total UC spend target since our simulation doesn't
    # distinguish the two components.
    uc_by_year: dict[int, float] = {}
    for row_num in range(6, 50):
        val = ws[f"B{row_num}"].value
        if val and str(val).strip().startswith("Universal credit"):
            values = _read_row(ws, row_num, _WELFARE_COL_TO_YEAR)
            for year, value in values.items():
                uc_by_year[year] = uc_by_year.get(year, 0.0) + value
    for year, value in uc_by_year.items():
        targets.append(
            {
                "name": f"obr/universal_credit_total/{year}",
                "variable": "universal_credit",
                "entity": "benunit",
                "aggregation": "sum",
                "filter": None,
                "value": value,
                "source": "obr",
                "year": year,
                "holdout": False,
            }
        )

    for label, name, variable, entity in benefit_rows:
        row = _find_row(ws, label)
        if row is None:
            continue
        values = _read_row(ws, row, _WELFARE_COL_TO_YEAR)
        for year, value in values.items():
            targets.append(
                {
                    "name": f"{name}/{year}",
                    "variable": variable,
                    "entity": entity,
                    "aggregation": "sum",
                    "filter": None,
                    "value": value,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

    wb.close()
    return targets


def _parse_council_tax() -> list[dict]:
    """Parse sheet 4.1 for council tax receipts."""
    wb = openpyxl.load_workbook(EXPENDITURE_FILE, data_only=True)
    ws = wb["4.1"]
    targets = []

    row = _find_row(ws, "Total net council tax receipts")
    if row:
        values = _read_row(ws, row, _CT_COL_TO_YEAR)
        for year, value in values.items():
            targets.append(
                {
                    "name": f"obr/council_tax/{year}",
                    "variable": "council_tax_annual",
                    "entity": "household",
                    "aggregation": "sum",
                    "filter": None,
                    "value": value,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

    wb.close()
    return targets


def _parse_fiscal_year(label: str) -> int | None:
    """Parse '2025-26' → 2025, or '2025/26' → 2025."""
    s = str(label).strip()
    for sep in ["-", "/"]:
        if sep in s:
            parts = s.split(sep)
            try:
                return int(parts[0])
            except ValueError:
                return None
    return None


def _read_fiscal_year_rows(
    ws, col_map: dict[str, str], max_row: int = 200
) -> list[tuple[int, dict[str, float]]]:
    """Scan column B for fiscal year labels (e.g. '2025-26') and read values.

    col_map maps a descriptive key to a column letter, e.g. {"employment": "C"}.
    Returns [(year, {key: value}), ...].
    """
    results = []
    for row in range(4, max_row):
        b = ws[f"B{row}"].value
        if b is None:
            continue
        year = _parse_fiscal_year(b)
        if year is None or year < 2020:
            continue
        vals = {}
        for key, col in col_map.items():
            v = ws[f"{col}{row}"].value
            if v is not None and isinstance(v, (int, float)):
                vals[key] = float(v)
        if vals:
            results.append((year, vals))
    return results


def _parse_economy() -> list[dict]:
    """Parse economy tables for labour market and income aggregates."""
    wb = openpyxl.load_workbook(ECONOMY_FILE, data_only=True)
    targets = []

    # ── 1.6 Labour market (fiscal year rows) ──
    ws = wb["1.6"]
    for year, vals in _read_fiscal_year_rows(
        ws,
        {
            "employment": "C",  # Employment 16+, millions
            "employees": "E",  # Employees 16+, millions
            "unemployment": "F",  # ILO unemployment, millions
            "total_hours": "J",  # Total hours worked, millions per week
            "comp_employees": "M",  # Compensation of employees, £bn
            "wages_salaries": "N",  # Wages and salaries, £bn
            "employer_social": "O",  # Employer social contributions, £bn
            "mixed_income": "P",  # Mixed income (self-employment), £bn
        },
    ):
        # Employment count: people with employment_income > 0
        if "employment" in vals:
            targets.append(
                {
                    "name": f"obr/employment_count/{year}",
                    "variable": "employment_income",
                    "entity": "person",
                    "aggregation": "count_nonzero",
                    "filter": None,
                    "value": vals["employment"] * 1e6,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

        # Total wages and salaries: sum of employment_income
        if "wages_salaries" in vals:
            targets.append(
                {
                    "name": f"obr/wages_salaries/{year}",
                    "variable": "employment_income",
                    "entity": "person",
                    "aggregation": "sum",
                    "filter": None,
                    "value": vals["wages_salaries"] * 1e9,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

        # Employer social contributions — skipped: OBR figure includes pensions
        # and other employer costs beyond NI. employer_ni already covered by
        # NI receipts target.

        # Mixed income ≈ total self-employment income
        if "mixed_income" in vals:
            targets.append(
                {
                    "name": f"obr/self_employment_income/{year}",
                    "variable": "self_employment_income",
                    "entity": "person",
                    "aggregation": "sum",
                    "filter": None,
                    "value": vals["mixed_income"] * 1e9,
                    "source": "obr",
                    "year": year,
                    "holdout": False,
                }
            )

        # Self-employment count
        if "mixed_income" in vals:
            targets.append(
                {
                    "name": f"obr/self_employed_count/{year}",
                    "variable": "self_employment_income",
                    "entity": "person",
                    "aggregation": "count_nonzero",
                    "filter": None,
                    "value": (vals["employment"] - vals.get("employees", 0)) * 1e6
                    if "employment" in vals and "employees" in vals
                    else 0,
                    "source": "obr",
                    "year": year,
                    "holdout": True,
                }
            )

        # Total hours worked — skipped: hours_worked not populated in EFRS.

    # RHDI (1.12) excluded — OBR national accounts definition differs from
    # HBAI net income (includes imputed rent, NPISH, etc.).
    # Housing stock (1.16) excluded — overlaps with ONS total_households.

    wb.close()
    return targets


def _backfill_2023(targets: list[dict]) -> list[dict]:
    """Back-extrapolate 2023 targets from 2024 outturn.

    The March 2026 EFO's earliest column is 2024/25 outturn. For 2023/24 we
    scale backwards using OBR growth rates: earnings growth for tax receipts,
    CPI for benefit spending, council tax growth for council tax.
    """
    # OBR growth rates for the 2023→2024 transition (from economy tables)
    EARNINGS_GROWTH_2024 = 0.0493
    CPI_GROWTH_2024 = 0.0253
    CT_GROWTH_2024 = 0.051

    # Which growth factor to use for each target prefix
    _DEFLATOR = {
        "obr/income_tax": EARNINGS_GROWTH_2024,
        "obr/ni_": EARNINGS_GROWTH_2024,
        "obr/vat_": EARNINGS_GROWTH_2024,
        "obr/fuel_duty": EARNINGS_GROWTH_2024,
        "obr/cgt_": EARNINGS_GROWTH_2024,
        "obr/sdlt_": EARNINGS_GROWTH_2024,
        "obr/council_tax": CT_GROWTH_2024,
        "obr/housing_benefit": CPI_GROWTH_2024,
        "obr/pip_dla": CPI_GROWTH_2024,
        "obr/attendance_allowance": CPI_GROWTH_2024,
        "obr/pension_credit": CPI_GROWTH_2024,
        "obr/carers_allowance": CPI_GROWTH_2024,
        "obr/child_benefit": CPI_GROWTH_2024,
        "obr/state_pension": CPI_GROWTH_2024,
        "obr/universal_credit": CPI_GROWTH_2024,
    }

    existing_2023 = {t["name"] for t in targets if t["year"] == 2023}
    extra = []
    for t in targets:
        if t["year"] != 2024 or t["source"] != "obr":
            continue
        name_2023 = t["name"].replace("/2024", "/2023")
        if name_2023 in existing_2023:
            continue
        # Find the right deflator
        growth = None
        for prefix, rate in _DEFLATOR.items():
            if t["name"].startswith(prefix):
                growth = rate
                break
        if growth is None:
            continue
        t2 = dict(t)
        t2["name"] = name_2023
        t2["year"] = 2023
        t2["value"] = t["value"] / (1 + growth)
        extra.append(t2)
    return targets + extra


def get_targets() -> list[dict]:
    targets = []
    if RECEIPTS_FILE.exists():
        targets.extend(_parse_receipts())
        targets.extend(_parse_it_nics_detail())
    if EXPENDITURE_FILE.exists():
        targets.extend(_parse_welfare())
        targets.extend(_parse_council_tax())
    if ECONOMY_FILE.exists():
        targets.extend(_parse_economy())
    targets = _backfill_2023(targets)
    return targets
