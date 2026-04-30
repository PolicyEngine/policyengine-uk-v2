"""Fetch DWP benefit statistics from the Stat-Xplore API and forecasts.

Queries caseloads for UC (with subgroup breakdowns), PIP, pension credit,
carer's allowance, attendance allowance, state pension, ESA, and DLA.
Results are cached locally to avoid repeated API calls.

The Stat-Xplore snapshot (latest month) is then scaled to all calibration
years (2024-2029) using DWP's own caseload forecasts from the Spring
Statement 2025 benefit expenditure and caseload tables.

Requires STAT_XPLORE_API_KEY environment variable to be set.
See: https://stat-xplore.dwp.gov.uk/webapi/online-help/Open-Data-API.html
"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path

import openpyxl
import requests

logger = logging.getLogger(__name__)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
CACHE_DIR = REPO_ROOT / "data" / "cache"
CACHE_FILE = CACHE_DIR / "dwp_stat_xplore.json"

API_BASE = "https://stat-xplore.dwp.gov.uk/webapi/rest/v1"
API_KEY = os.environ.get("STAT_XPLORE_API_KEY", "")


def _headers() -> dict:
    return {"apiKey": API_KEY, "Content-Type": "application/json"}


def _query_table(
    database: str,
    measures: list[str],
    dimensions: list[list[str]],
) -> dict:
    """Send a table query to stat-xplore and return the JSON response."""
    payload: dict = {
        "database": database,
        "measures": measures,
        "dimensions": dimensions,
    }
    r = requests.post(f"{API_BASE}/table", headers=_headers(), json=payload, timeout=30)
    r.raise_for_status()
    return r.json()


def _extract_year(result: dict) -> int:
    """Extract the year from the auto-selected date field."""
    for field in result.get("fields", []):
        for item in field.get("items", []):
            for label in item.get("labels", []):
                for part in str(label).replace("-", " ").split():
                    if part.isdigit():
                        y = int(part)
                        return y if y > 100 else 2000 + y
    return 2025


def _extract_total(result: dict) -> float | None:
    """Extract the single value from a no-dimension query."""
    cubes = result.get("cubes", {})
    if not cubes:
        return None
    values = next(iter(cubes.values()))["values"]
    # Unwrap nested lists (stat-xplore wraps in [date][value])
    while isinstance(values, list) and len(values) == 1:
        values = values[0]
    return values if isinstance(values, (int, float)) else None


def _extract_breakdown(result: dict) -> list[tuple[str, float]]:
    """Extract label/value pairs from a single-dimension query.

    Stat-xplore auto-adds the date dimension, so the response has two fields:
    date (1 item = latest month) and the requested dimension (N items).
    Values are shaped [1][N].
    """
    fields = result.get("fields", [])
    cubes = result.get("cubes", {})
    if not cubes:
        return []
    vals = next(iter(cubes.values()))["values"]

    # Find the non-date dimension
    dim_field = None
    for f in fields:
        if "month" not in f["label"].lower() and "date" not in f["label"].lower():
            dim_field = f
            break
    if dim_field is None:
        return []

    items = dim_field["items"]
    # Values are [date_idx][dim_idx] — take last date row
    row = vals[-1] if isinstance(vals[0], list) else vals
    pairs = []
    for i, item in enumerate(items):
        v = row[i] if isinstance(row, list) else row
        if v is not None and v > 0:
            pairs.append((item["labels"][0], float(v)))
    return pairs


# ── Simple total caseload queries ──────────────────────────────────────────


# (database, measure, target_name, survey_variable, entity)
_SIMPLE_BENEFITS = [
    (
        "str:database:UC_Monthly",
        "str:count:UC_Monthly:V_F_UC_CASELOAD_FULL",
        "dwp/uc_total_claimants",
        "universal_credit",
        "person",
    ),
    (
        "str:database:PIP_Monthly_new",
        "str:count:PIP_Monthly_new:V_F_PIP_MONTHLY",
        "dwp/pip_total_claimants",
        "pip_daily_living",
        "person",
    ),
    (
        "str:database:PC_New",
        "str:count:PC_New:V_F_PC_CASELOAD_New",
        "dwp/pension_credit_claimants",
        "pension_credit",
        "person",
    ),
    (
        "str:database:CA_In_Payment_New",
        "str:count:CA_In_Payment_New:V_F_CA_In_Payment_New",
        "dwp/carers_allowance_claimants",
        "carers_allowance",
        "person",
    ),
    (
        "str:database:AA_In_Payment_New",
        "str:count:AA_In_Payment_New:V_F_AA_In_Payment_New",
        "dwp/attendance_allowance_claimants",
        "attendance_allowance",
        "person",
    ),
    (
        "str:database:SP_New",
        "str:count:SP_New:V_F_SP_CASELOAD_New",
        "dwp/state_pension_claimants",
        "state_pension",
        "person",
    ),
    (
        "str:database:ESA_Caseload_new",
        "str:count:ESA_Caseload_new:V_F_ESA_NEW",
        "dwp/esa_claimants",
        "esa_income",
        "person",
    ),
    (
        "str:database:DLA_In_Payment_New",
        "str:count:DLA_In_Payment_New:V_F_DLA_In_Payment_New",
        "dwp/dla_claimants",
        "dla_care",
        "person",
    ),
]


def _fetch_simple_benefits() -> list[dict]:
    """Fetch total caseload for each benefit."""
    targets = []
    for database, measure, name, variable, entity in _SIMPLE_BENEFITS:
        try:
            result = _query_table(database, [measure], [])
            total = _extract_total(result)
            if total is not None:
                year = _extract_year(result)
                targets.append(
                    {
                        "name": name,
                        "variable": variable,
                        "entity": entity,
                        "aggregation": "count_nonzero",
                        "filter": None,
                        "value": total,
                        "source": "dwp",
                        "year": year,
                        "holdout": False,
                    }
                )
        except Exception as e:
            logger.warning("Failed to fetch %s: %s", name, e)
    return targets


# ── UC subgroup breakdowns (households) ────────────────────────────────────

_UC_HH_DB = "str:database:UC_Households"
_UC_HH_COUNT = "str:count:UC_Households:V_F_UC_HOUSEHOLDS"
_UC_HH_FIELD = "str:field:UC_Households:V_F_UC_HOUSEHOLDS"


def _fetch_uc_breakdowns() -> list[dict]:
    """Fetch UC household breakdowns by family type, entitlement elements, etc."""
    targets = []

    # UC households by family type — map to benunit_filter conditions
    try:
        result = _query_table(
            _UC_HH_DB,
            [_UC_HH_COUNT],
            [[f"{_UC_HH_FIELD}:hnfamily_type"]],
        )
        year = _extract_year(result)
        for label, value in _extract_breakdown(result):
            slug = label.lower().replace(",", "").replace(" ", "_")
            if "unknown" in slug or "missing" in slug:
                continue
            # Map family type labels to benunit filter conditions
            bf = {}
            if "single" in slug and "no_child" in slug:
                bf = {"is_couple": False, "has_children": False}
            elif "single" in slug and "child" in slug:
                bf = {"is_couple": False, "has_children": True}
            elif "couple" in slug and "no_child" in slug:
                bf = {"is_couple": True, "has_children": False}
            elif "couple" in slug and "child" in slug:
                bf = {"is_couple": True, "has_children": True}

            targets.append(
                {
                    "name": f"dwp/uc_households_{slug}",
                    "variable": "universal_credit",
                    "entity": "benunit",
                    "aggregation": "count_nonzero",
                    "filter": None,
                    "benunit_filter": bf if bf else None,
                    "value": value,
                    "source": "dwp",
                    "year": year,
                    "holdout": True,
                }
            )
    except Exception as e:
        logger.warning("Failed to fetch UC family type breakdown: %s", e)

    # UC households with child entitlement
    try:
        result = _query_table(
            _UC_HH_DB,
            [_UC_HH_COUNT],
            [[f"{_UC_HH_FIELD}:HCCHILD_ENTITLEMENT"]],
        )
        year = _extract_year(result)
        for label, value in _extract_breakdown(result):
            if label.lower() == "yes":
                targets.append(
                    {
                        "name": "dwp/uc_households_with_children",
                        "variable": "universal_credit",
                        "entity": "benunit",
                        "aggregation": "count_nonzero",
                        "filter": None,
                        "benunit_filter": {"has_children": True},
                        "value": value,
                        "source": "dwp",
                        "year": year,
                        "holdout": False,
                    }
                )
    except Exception as e:
        logger.warning("Failed to fetch UC child entitlement breakdown: %s", e)

    # UC households with LCWRA entitlement (disability element)
    try:
        result = _query_table(
            _UC_HH_DB,
            [_UC_HH_COUNT],
            [[f"{_UC_HH_FIELD}:HCLCW_ENTITLEMENT"]],
        )
        year = _extract_year(result)
        for label, value in _extract_breakdown(result):
            slug = label.lower().replace(" ", "_").replace("/", "_")
            if slug == "lcwra":
                targets.append(
                    {
                        "name": "dwp/uc_households_lcwra",
                        "variable": "universal_credit",
                        "entity": "benunit",
                        "aggregation": "count_nonzero",
                        "filter": None,
                        "benunit_filter": {"has_lcwra": True},
                        "value": value,
                        "source": "dwp",
                        "year": year,
                        "holdout": False,
                    }
                )
            elif slug == "lcw":
                targets.append(
                    {
                        "name": "dwp/uc_households_lcw",
                        "variable": "universal_credit",
                        "entity": "benunit",
                        "aggregation": "count_nonzero",
                        "filter": None,
                        "benunit_filter": {"has_lcw": True},
                        "value": value,
                        "source": "dwp",
                        "year": year,
                        "holdout": True,
                    }
                )
    except Exception as e:
        logger.warning("Failed to fetch UC LCW breakdown: %s", e)

    # UC households with carer entitlement
    try:
        result = _query_table(
            _UC_HH_DB,
            [_UC_HH_COUNT],
            [[f"{_UC_HH_FIELD}:HCCARER_ENTITLEMENT"]],
        )
        year = _extract_year(result)
        for label, value in _extract_breakdown(result):
            if label.lower() == "yes":
                targets.append(
                    {
                        "name": "dwp/uc_households_with_carer",
                        "variable": "universal_credit",
                        "entity": "benunit",
                        "aggregation": "count_nonzero",
                        "filter": None,
                        "benunit_filter": {"has_carer": True},
                        "value": value,
                        "source": "dwp",
                        "year": year,
                        "holdout": True,
                    }
                )
    except Exception as e:
        logger.warning("Failed to fetch UC carer breakdown: %s", e)

    # UC households with housing entitlement
    try:
        result = _query_table(
            _UC_HH_DB,
            [_UC_HH_COUNT],
            [[f"{_UC_HH_FIELD}:TENURE"]],
        )
        year = _extract_year(result)
        for label, value in _extract_breakdown(result):
            if label.lower() == "yes":
                targets.append(
                    {
                        "name": "dwp/uc_households_with_housing",
                        "variable": "universal_credit",
                        "entity": "benunit",
                        "aggregation": "count_nonzero",
                        "filter": None,
                        "benunit_filter": {"has_housing": True},
                        "value": value,
                        "source": "dwp",
                        "year": year,
                        "holdout": False,
                    }
                )
    except Exception as e:
        logger.warning("Failed to fetch UC housing breakdown: %s", e)

    # UC households by monthly payment band — constrains the UC amount distribution
    try:
        result = _query_table(
            _UC_HH_DB,
            [_UC_HH_COUNT],
            [[f"{_UC_HH_FIELD}:hnpayment_band"]],
        )
        year = _extract_year(result)
        pairs = _extract_breakdown(result)

        # Consolidate into wider bands (monthly £ → annual £ for filter ranges).
        # Stat-xplore bands are £100-wide from £0 to £2500+. We group into ~£300-400
        # bands to keep the target count reasonable while constraining the distribution.
        _BAND_GROUPS = [
            ("0_to_300", 0, 300),
            ("300_to_600", 300, 600),
            ("600_to_900", 600, 900),
            ("900_to_1200", 900, 1200),
            ("1200_to_1500", 1200, 1500),
            ("1500_to_2000", 1500, 2000),
            ("2000_plus", 2000, 999999),
        ]

        # Parse stat-xplore band labels into (lower_monthly, upper_monthly, count)
        parsed_bands = []
        for label, count in pairs:
            low_label = label.lower().strip()
            if "no payment" in low_label:
                parsed_bands.append((0, 0, count))
            elif "or over" in low_label:
                # e.g. "£2500.01 or over"
                val = float(low_label.split("£")[1].split(" ")[0])
                parsed_bands.append((val, 999999, count))
            elif " to " in low_label:
                parts = low_label.replace("£", "").replace(",", "").split(" to ")
                lo = float(parts[0])
                hi = float(parts[1])
                parsed_bands.append((lo, hi, count))

        # Aggregate into grouped bands
        for group_name, group_lo, group_hi in _BAND_GROUPS:
            group_count = 0.0
            for lo, hi, count in parsed_bands:
                if lo == 0 and hi == 0:
                    continue  # skip "no payment"
                band_mid = (lo + min(hi, 5000)) / 2.0
                if group_lo <= band_mid < group_hi:
                    group_count += count

            if group_count > 0:
                # Filter range: convert monthly band to annual
                annual_lo = group_lo * 12.0
                annual_hi = group_hi * 12.0

                targets.append(
                    {
                        "name": f"dwp/uc_payment_band_{group_name}",
                        "variable": "universal_credit",
                        "entity": "benunit",
                        "aggregation": "count_nonzero",
                        "filter": {
                            "variable": "universal_credit",
                            "min": annual_lo,
                            "max": annual_hi,
                        },
                        "value": group_count,
                        "source": "dwp",
                        "year": year,
                        "holdout": False,
                    }
                )

    except Exception as e:
        logger.warning("Failed to fetch UC payment band breakdown: %s", e)

    return targets


DWP_FORECAST_URL = (
    "https://assets.publishing.service.gov.uk/media/68f8923724fc2bb7eed11ac8/"
    "outturn-and-forecast-tables-spring-statement-2025.xlsx"
)
DWP_FORECAST_FILE = CACHE_DIR / "dwp_spring_statement_2025.xlsx"

CALIBRATION_YEARS = range(2023, 2030)  # 2023/24 through 2029/30

# Column 80 = 2024/25, ..., 85 = 2029/30 in the DWP forecast xlsx
_FORECAST_COL_TO_YEAR = {79: 2023, 80: 2024, 81: 2025, 82: 2026, 83: 2027, 84: 2028, 85: 2029}


def _download_forecast() -> Path:
    """Download the DWP forecast xlsx if not cached."""
    if DWP_FORECAST_FILE.exists():
        return DWP_FORECAST_FILE
    logger.info("Downloading DWP forecast tables...")
    r = requests.get(DWP_FORECAST_URL, timeout=60, allow_redirects=True)
    r.raise_for_status()
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    DWP_FORECAST_FILE.write_bytes(r.content)
    return DWP_FORECAST_FILE


def _find_forecast_row(ws, label: str, start_row: int = 1, max_row: int = 200) -> int | None:
    """Find the first row in column B starting with label."""
    for row in range(start_row, max_row + 1):
        val = ws.cell(row=row, column=2).value
        if val and str(val).strip().startswith(label):
            return row
    return None


def _read_forecast_row(ws, row: int) -> dict[int, float]:
    """Read caseload values (thousands) from a forecast row."""
    result = {}
    for col, year in _FORECAST_COL_TO_YEAR.items():
        val = ws.cell(row=row, column=col).value
        if val is not None and isinstance(val, (int, float)):
            result[year] = float(val) * 1e3  # thousands → people
    return result


def _parse_caseload_forecasts() -> dict[str, dict[int, float]]:
    """Parse DWP forecast xlsx for benefit caseload projections.

    Returns {benefit_key: {year: caseload}} for each benefit.
    """
    try:
        path = _download_forecast()
    except Exception as e:
        logger.warning("Failed to download DWP forecast: %s", e)
        return {}

    wb = openpyxl.load_workbook(path, data_only=True)
    forecasts: dict[str, dict[int, float]] = {}

    # UC caseloads from "Universal Credit and equivalent" sheet
    ws = wb["Universal Credit and equivalent"]
    uc_row = _find_forecast_row(ws, "Universal Credit", start_row=48)
    if uc_row:
        forecasts["universal_credit"] = _read_forecast_row(ws, uc_row)

    uc_carer_row = _find_forecast_row(ws, "Universal Credit Carers Element", start_row=48)
    if uc_carer_row:
        forecasts["uc_carer_element"] = _read_forecast_row(ws, uc_carer_row)

    uc_housing_row = _find_forecast_row(ws, "Universal Credit Housing Element", start_row=48)
    if uc_housing_row:
        forecasts["uc_housing_element"] = _read_forecast_row(ws, uc_housing_row)

    # LCWRA from health element breakdown
    lcwra_row = _find_forecast_row(ws, "of which limited capability for work and work-related activi", start_row=48)
    if lcwra_row:
        forecasts["uc_lcwra"] = _read_forecast_row(ws, lcwra_row)

    lcw_row = _find_forecast_row(ws, "of which limited capability for work", start_row=48)
    if lcw_row:
        # Make sure we didn't pick up the LCWRA row
        label = str(ws.cell(row=lcw_row, column=2).value).strip()
        if "related" not in label:
            forecasts["uc_lcw"] = _read_forecast_row(ws, lcw_row)

    esa_row = _find_forecast_row(ws, "Employment and Support Allowance", start_row=48)
    if esa_row:
        forecasts["esa"] = _read_forecast_row(ws, esa_row)

    # Disability benefits sheet
    ws = wb["Disability benefits"]
    pip_row = _find_forecast_row(ws, "Personal Independence Payment", start_row=50)
    if pip_row:
        forecasts["pip"] = _read_forecast_row(ws, pip_row)

    dla_row = _find_forecast_row(ws, "Disability Living Allowance", start_row=50)
    if dla_row:
        forecasts["dla"] = _read_forecast_row(ws, dla_row)

    aa_row = _find_forecast_row(ws, "Attendance Allowance", start_row=50)
    if aa_row:
        forecasts["attendance_allowance"] = _read_forecast_row(ws, aa_row)

    # Carer's Allowance sheet
    ws = wb["Carers Allowance"]
    ca_total_row = _find_forecast_row(ws, "Total", start_row=14)
    if ca_total_row:
        forecasts["carers_allowance"] = _read_forecast_row(ws, ca_total_row)

    # Pension Credit sheet
    ws = wb["Pension Credit"]
    pc_row = _find_forecast_row(ws, "Total Pension Credit", start_row=18)
    if pc_row:
        forecasts["pension_credit"] = _read_forecast_row(ws, pc_row)

    # State Pension sheet
    ws = wb["State Pension"]
    sp_row = _find_forecast_row(ws, "Total State Pension Caseload", start_row=28)
    if sp_row:
        forecasts["state_pension"] = _read_forecast_row(ws, sp_row)

    wb.close()
    return forecasts


def _scale_targets_to_years(
    base_targets: list[dict],
    forecasts: dict[str, dict[int, float]],
) -> list[dict]:
    """Scale stat-xplore snapshot targets to all calibration years using DWP forecasts.

    For each base target (from stat-xplore, typically 2025), compute a scaling
    factor from the DWP forecast caseload trajectory and emit a target for each year.
    """
    # Map target names to forecast keys for scaling
    _FORECAST_KEY = {
        "dwp/uc_total_claimants": "universal_credit",
        "dwp/pip_total_claimants": "pip",
        "dwp/pension_credit_claimants": "pension_credit",
        "dwp/carers_allowance_claimants": "carers_allowance",
        "dwp/attendance_allowance_claimants": "attendance_allowance",
        "dwp/state_pension_claimants": "state_pension",
        "dwp/esa_claimants": "esa",
        "dwp/dla_claimants": "dla",
        "dwp/uc_households_with_children": "universal_credit",
        "dwp/uc_households_lcwra": "uc_lcwra",
        "dwp/uc_households_lcw": "uc_lcw",
        "dwp/uc_households_with_carer": "uc_carer_element",
        "dwp/uc_households_with_housing": "uc_housing_element",
        # Family type breakdowns scale with total UC
        "dwp/uc_households_single_no_children": "universal_credit",
        "dwp/uc_households_single_with_children": "universal_credit",
        "dwp/uc_households_couple_no_children": "universal_credit",
        "dwp/uc_households_couple_with_children": "universal_credit",
        # Payment band breakdowns scale with total UC
        "dwp/uc_payment_band_0_to_300": "universal_credit",
        "dwp/uc_payment_band_300_to_600": "universal_credit",
        "dwp/uc_payment_band_600_to_900": "universal_credit",
        "dwp/uc_payment_band_900_to_1200": "universal_credit",
        "dwp/uc_payment_band_1200_to_1500": "universal_credit",
        "dwp/uc_payment_band_1500_to_2000": "universal_credit",
        "dwp/uc_payment_band_2000_plus": "universal_credit",
        # Age band breakdowns scale with total UC
        "dwp/uc_age_16_24": "universal_credit",
        "dwp/uc_age_25_34": "universal_credit",
        "dwp/uc_age_35_49": "universal_credit",
        "dwp/uc_age_50_64": "universal_credit",
        "dwp/uc_age_65_plus": "universal_credit",
    }

    scaled: list[dict] = []
    for target in base_targets:
        base_year = target["year"]
        forecast_key = _FORECAST_KEY.get(target["name"])
        forecast_series = forecasts.get(forecast_key, {}) if forecast_key else {}
        base_forecast = forecast_series.get(base_year, 0)

        for year in CALIBRATION_YEARS:
            year_forecast = forecast_series.get(year, 0)
            if base_forecast > 0 and year_forecast > 0:
                scale = year_forecast / base_forecast
            else:
                scale = 1.0

            t = dict(target)
            t["name"] = f"{target['name']}/{year}"
            t["year"] = year
            t["value"] = target["value"] * scale
            scaled.append(t)

    return scaled


def get_targets() -> list[dict]:
    if CACHE_FILE.exists():
        logger.info("Using cached DWP targets: %s", CACHE_FILE)
        base_targets = json.loads(CACHE_FILE.read_text())
    elif API_KEY:
        base_targets = []
        base_targets.extend(_fetch_simple_benefits())
        base_targets.extend(_fetch_uc_breakdowns())
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        CACHE_FILE.write_text(json.dumps(base_targets, indent=2))
        logger.info("Cached %d DWP base targets to %s", len(base_targets), CACHE_FILE)
    else:
        logger.warning(
            "STAT_XPLORE_API_KEY not set and no cache — skipping DWP targets. "
            "Set the env var and re-run to fetch from stat-xplore."
        )
        return []

    # Parse DWP caseload forecasts and scale base targets to all years
    forecasts = _parse_caseload_forecasts()
    if forecasts:
        return _scale_targets_to_years(base_targets, forecasts)

    # Fallback: emit base targets as-is (single year only)
    logger.warning("No DWP forecasts available — emitting base targets for single year only")
    return base_targets
