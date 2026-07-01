"""Generate parameter-YAML growth_factors from the OBR Economic and Fiscal Outlook.

Reads the OBR EFO "detailed forecast tables: economy" workbook (fiscal-year rows)
and writes the `growth_factors:` block of each `parameters/YYYY_YY.yaml`:

  cpi_rate                ← Table 1.7 "CPI" (year-on-year, fiscal year)
  gdp_deflator            ← Table 1.7 "GDP deflator"
  earnings_growth         ← Table 1.6 "Average weekly earnings growth"
  gdp_pc_growth           ← Table 1.4 nominal GDP (true fiscal-year sums from
                            quarterly rows — the workbook's own "fiscal year"
                            block in 1.4 is centred on end-March, i.e. Q4..Q3,
                            and must NOT be used) ÷ 16+ population
  mixed_income_pc_growth  ← Table 1.6 "Mixed income" ÷ 16+ population
  savings_interest_growth ← Table 1.9 "Deposit rate" (growth in the level)
  rent_growth             ← Table 1.7 "Actual rents for housing" (YoY)
  population_growth       ← 16+ population from Table 1.6 (employment ÷
                            employment rate)
  council_tax_growth      ← MHCLG "Council Tax levels set by local authorities
                            in England", Table 3 average Band D percentage
                            change (outturn to 2026); OBR-derived forecast
                            values for later years

The cpi_rate/gdp_deflator/earnings_growth trio is only rewritten from
`--from-year` (default 2025) — outturn years carry values from their own
historical releases. The six newer indices are written for ALL years the
sources cover, since no earlier vintage of them exists in the YAMLs.

The OBR workbook is the canonical source; it lives in the microdata bucket at
`gs://policyengine-uk-microdata/obr/efo_2026_03_economy.xlsx` and is fetched on
run (cached locally). The MHCLG council tax ODS is cached alongside it.

Usage::

    uv run python data/gen_growth_factors.py --dry-run     # show diff, write nothing
    uv run python data/gen_growth_factors.py               # rewrite YAMLs
    uv run python data/gen_growth_factors.py --from-year 2025
"""

from __future__ import annotations

import argparse
import re
import subprocess
from pathlib import Path

import openpyxl
from rich.console import Console
from rich.table import Table

REPO_ROOT = Path(__file__).resolve().parent.parent
PARAMS_DIR = REPO_ROOT / "parameters"

GCS_OBJECT = "gs://policyengine-uk-microdata/obr/efo_2026_03_economy.xlsx"
LOCAL_CACHE = Path.home() / ".policyengine-uk-data" / "obr" / "efo_2026_03_economy.xlsx"
SOURCE_LABEL = "OBR Economic and Fiscal Outlook, March 2026"

# MHCLG average Band D council tax (England), Table 3 of "Council Tax levels
# set by local authorities in England 2026 to 2027" (accredited official
# statistics, published 25 March 2026).
MHCLG_CT_URL = (
    "https://assets.publishing.service.gov.uk/media/"
    "69de1fa63e81003ae0422508/Tables_1-9_2026-27.ods"
)
MHCLG_CT_CACHE = (
    Path.home()
    / ".policyengine-uk-data"
    / "mhclg"
    / "council_tax_tables_1-9_2026-27.ods"
)
MHCLG_CT_LABEL = "MHCLG council tax levels 2026-27, Table 3 (England average Band D)"

# Council tax growth for years beyond the MHCLG outturn series: OBR-derived
# forecast values carried over from the pre-existing uprating table
# (src/data/mod.rs yoy_rates, OBR EFO November 2025 vintage).
COUNCIL_TAX_FORECAST = {2027: 0.0579, 2028: 0.0565, 2029: 0.0547, 2030: 0.0542}
COUNCIL_TAX_FORECAST_LABEL = "OBR EFO November 2025 (carried over)"

# First fiscal year the OBR EFO treats as a forecast (data is outturn before this).
DEFAULT_FROM_YEAR = 2025

# First fiscal year for which the extended (non-cpi/deflator/earnings) indices
# are written. 2017 covers every transition needed to de-uprate the 2024-based
# EFRS panel back to 2016.
EXTENDED_FROM_YEAR = 2017

# Sheet / column locations, verified against the March 2026 workbook.
INFLATION_SHEET = "1.7"
INFLATION_CPI_COL = 4
INFLATION_RENT_COL = 8  # "Actual rents for housing", YoY %
INFLATION_GDP_DEFLATOR_COL = 11
LABOUR_SHEET = "1.6"
LABOUR_AWE_GROWTH_HEADER = "Average weekly earnings growth"
LABOUR_EMPLOYMENT_COL = 2  # employment 16+, millions
LABOUR_EMPLOYMENT_RATE_COL = 3  # employment rate 16+, per cent
LABOUR_MIXED_INCOME_COL = 15  # mixed income, £bn
NOMINAL_GDP_SHEET = "1.4"
NOMINAL_GDP_COL = 2  # £bn, quarterly rows (NSA)
MARKET_SHEET = "1.9"
MARKET_DEPOSIT_RATE_COL = 5  # deposit rate, per cent

console = Console()

_FISCAL_RE = re.compile(r"^(\d{4})-(\d{2})$")
_QUARTER_RE = re.compile(r"^(\d{4})Q([1-4])$")


def ensure_workbook() -> Path:
    """Download the OBR workbook from GCS if not cached locally."""
    if LOCAL_CACHE.exists():
        return LOCAL_CACHE
    LOCAL_CACHE.parent.mkdir(parents=True, exist_ok=True)
    console.print(f"  downloading {GCS_OBJECT}")
    subprocess.run(
        ["gcloud", "storage", "cp", GCS_OBJECT, str(LOCAL_CACHE)], check=True
    )
    return LOCAL_CACHE


def _fiscal_rows(ws) -> dict[int, tuple]:
    """Map the entered fiscal year (e.g. 2026 for '2026-27') to its row tuple."""
    out: dict[int, tuple] = {}
    for r in ws.iter_rows(values_only=True):
        label = r[1] if len(r) > 1 else None
        if isinstance(label, str):
            m = _FISCAL_RE.match(label.strip())
            if m:
                out[int(m.group(1))] = r
    return out


def _awe_growth_col(ws) -> int:
    header = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    for j, h in enumerate(header):
        if isinstance(h, str) and LABOUR_AWE_GROWTH_HEADER in h:
            return j
    raise RuntimeError(
        f"Could not find '{LABOUR_AWE_GROWTH_HEADER}' column in sheet {LABOUR_SHEET}"
    )


def _quarterly_values(ws, col: int) -> dict[tuple[int, int], float]:
    """Map (year, quarter) to the value in `col` for quarterly rows."""
    out: dict[tuple[int, int], float] = {}
    for r in ws.iter_rows(values_only=True):
        label = r[1] if len(r) > 1 else None
        if isinstance(label, str):
            m = _QUARTER_RE.match(label.strip())
            if m and len(r) > col and r[col] is not None:
                out[(int(m.group(1)), int(m.group(2)))] = float(r[col])
    return out


def _fy_sum_from_quarters(q: dict[tuple[int, int], float], year: int) -> float | None:
    """True April–March fiscal-year sum: Q2..Q4 of `year` plus Q1 of `year`+1.

    Table 1.4's own annual block is labelled "Centred end-March" (Q4..Q3), so
    fiscal-year GDP must be built from the quarterly rows instead.
    """
    keys = [(year, 2), (year, 3), (year, 4), (year + 1, 1)]
    if any(k not in q for k in keys):
        return None
    return sum(q[k] for k in keys)


def read_obr_rates(xlsx: Path) -> dict[int, dict[str, float]]:
    """Return {fiscal_year: {field: rate}} (as fractions) from the EFO workbook.

    Growth rates are keyed on the fiscal year they enter (the transition from
    year-1 to year), matching the convention of the uprating tables.
    """
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    infl = _fiscal_rows(wb[INFLATION_SHEET])
    lab_ws = wb[LABOUR_SHEET]
    lab = _fiscal_rows(lab_ws)
    awe_col = _awe_growth_col(lab_ws)
    market = _fiscal_rows(wb[MARKET_SHEET])
    gdp_q = _quarterly_values(wb[NOMINAL_GDP_SHEET], NOMINAL_GDP_COL)

    # Fiscal-year level series (verified true April–March aggregates, unlike 1.4's
    # centred block): 16+ population, mixed income per head, nominal GDP per head,
    # deposit rate.
    pop16: dict[int, float] = {}
    mixed_pc: dict[int, float] = {}
    gdp_pc: dict[int, float] = {}
    deposit: dict[int, float] = {}
    for year, row in lab.items():
        emp = row[LABOUR_EMPLOYMENT_COL]
        emp_rate = row[LABOUR_EMPLOYMENT_RATE_COL]
        mixed = row[LABOUR_MIXED_INCOME_COL]
        if emp is None or emp_rate is None:
            continue
        pop16[year] = float(emp) / (float(emp_rate) / 100.0)
        if mixed is not None:
            mixed_pc[year] = float(mixed) / pop16[year]
        gdp = _fy_sum_from_quarters(gdp_q, year)
        if gdp is not None:
            gdp_pc[year] = gdp / pop16[year]
    for year, row in market.items():
        if (
            len(row) > MARKET_DEPOSIT_RATE_COL
            and row[MARKET_DEPOSIT_RATE_COL] is not None
        ):
            deposit[year] = float(row[MARKET_DEPOSIT_RATE_COL])

    def growth(series: dict[int, float], year: int) -> float | None:
        if year in series and (year - 1) in series and series[year - 1]:
            return series[year] / series[year - 1] - 1.0
        return None

    rates: dict[int, dict[str, float]] = {}
    for year in sorted(set(infl) & set(lab)):
        row: dict[str, float] = {}
        cpi = infl[year][INFLATION_CPI_COL]
        gdp_defl = infl[year][INFLATION_GDP_DEFLATOR_COL]
        awe = lab[year][awe_col]
        rent = infl[year][INFLATION_RENT_COL]
        if cpi is not None:
            row["cpi_rate"] = cpi / 100.0
        if gdp_defl is not None:
            row["gdp_deflator"] = gdp_defl / 100.0
        if awe is not None:
            row["earnings_growth"] = awe / 100.0
        if rent is not None:
            row["rent_growth"] = rent / 100.0
        for field, series in [
            ("gdp_pc_growth", gdp_pc),
            ("mixed_income_pc_growth", mixed_pc),
            ("savings_interest_growth", deposit),
            ("population_growth", pop16),
        ]:
            g = growth(series, year)
            if g is not None:
                row[field] = g
        if row:
            rates[year] = row
    return rates


def ensure_mhclg_workbook() -> Path:
    """Download the MHCLG council tax ODS if not cached locally."""
    if MHCLG_CT_CACHE.exists():
        return MHCLG_CT_CACHE
    MHCLG_CT_CACHE.parent.mkdir(parents=True, exist_ok=True)
    console.print(f"  downloading {MHCLG_CT_URL}")
    subprocess.run(["curl", "-sL", MHCLG_CT_URL, "-o", str(MHCLG_CT_CACHE)], check=True)
    return MHCLG_CT_CACHE


def read_council_tax_rates(ods: Path) -> dict[int, float]:
    """Return {fiscal_year: growth} for England average Band D council tax.

    Table 3 rows are labelled "2016 to 2017" etc.; the percentage change column
    is the change *into* the first-named year's fiscal year (e.g. "2017 to 2018"
    is the transition into FY 2017). Forecast years beyond the MHCLG series come
    from COUNCIL_TAX_FORECAST.
    """
    import pandas as pd

    df = pd.read_excel(ods, engine="odf", sheet_name="Table_3", header=None)
    label_re = re.compile(r"^(\d{4}) to (\d{4})")
    out: dict[int, float] = {}
    for _, r in df.iterrows():
        m = label_re.match(str(r[0]).strip())
        if m and pd.notna(r[2]):
            out[int(m.group(1))] = float(r[2]) / 100.0
    out.update(COUNCIL_TAX_FORECAST)
    return out


def _yaml_path(year: int) -> Path:
    return PARAMS_DIR / f"{year}_{(year + 1) % 100:02d}.yaml"


# Field → (source comment, whether it belongs to the legacy trio rewritten only
# from --from-year). Order here is the order rendered into the YAML.
_FIELDS: dict[str, tuple[str, bool]] = {
    "cpi_rate": ("CPI inflation, EFO Table 1.7", True),
    "gdp_deflator": ("GDP deflator, EFO Table 1.7", True),
    "earnings_growth": ("average weekly earnings growth, EFO Table 1.6", True),
    "gdp_pc_growth": (
        "nominal GDP per 16+ head, EFO Tables 1.4 (quarterly) + 1.6",
        False,
    ),
    "mixed_income_pc_growth": ("mixed income per 16+ head, EFO Table 1.6", False),
    "savings_interest_growth": ("deposit rate growth, EFO Table 1.9", False),
    "rent_growth": ("actual rents for housing, EFO Table 1.7", False),
    "population_growth": ("16+ population, EFO Table 1.6", False),
    "council_tax_growth": ("England average Band D, MHCLG Table 3", False),
}


def _render_block(year: int, fields: dict[str, float]) -> str:
    fy = f"{year}/{(year + 1) % 100:02d}"
    width = max(
        len(f"{name}: {fields[name]:.5f}") for name in _FIELDS if name in fields
    )
    lines = [
        "growth_factors:",
        f"  # {SOURCE_LABEL}; council tax: {MHCLG_CT_LABEL}",
        f"  # Rates are the year-on-year change entering FY {fy}.",
    ]
    for name, (comment, _) in _FIELDS.items():
        if name in fields:
            if name == "council_tax_growth" and year in COUNCIL_TAX_FORECAST:
                comment = f"England average Band D, {COUNCIL_TAX_FORECAST_LABEL}"
            entry = f"{name}: {fields[name]:.5f}"
            lines.append(f"  {entry:<{width}}  # {fields[name] * 100:+.3f}% {comment}")
    return "\n".join(lines) + "\n"


def _read_existing(year: int) -> dict[str, float]:
    path = _yaml_path(year)
    if not path.exists():
        return {}
    import yaml

    with path.open() as f:
        data = yaml.safe_load(f)
    return (data or {}).get("growth_factors") or {}


def _rewrite_block(year: int, block: str) -> bool:
    """Replace the trailing growth_factors block in the YAML. Returns True if written."""
    path = _yaml_path(year)
    if not path.exists():
        return False
    text = path.read_text()
    idx = text.find("\ngrowth_factors:")
    if idx == -1:
        # No existing block: append after a blank line.
        new_text = text.rstrip("\n") + "\n\n" + block
    else:
        new_text = text[: idx + 1] + block
    if not new_text.endswith("\n"):
        new_text += "\n"
    path.write_text(new_text)
    return True


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Generate growth_factors YAML blocks from OBR EFO tables."
    )
    ap.add_argument(
        "--dry-run", action="store_true", help="Show the diff; write nothing."
    )
    ap.add_argument(
        "--from-year",
        type=int,
        default=DEFAULT_FROM_YEAR,
        help="First fiscal year whose cpi/deflator/earnings values are rewritten "
        f"from the workbook (default {DEFAULT_FROM_YEAR}); before this they are "
        "preserved from the existing YAML. The extended indices are written from "
        f"{EXTENDED_FROM_YEAR} regardless.",
    )
    ap.add_argument(
        "--to-year",
        type=int,
        default=None,
        help="Last fiscal year to write (default: all available).",
    )
    args = ap.parse_args()

    xlsx = ensure_workbook()
    rates = read_obr_rates(xlsx)
    council_tax = read_council_tax_rates(ensure_mhclg_workbook())

    years = [
        y
        for y in sorted(rates)
        if y >= EXTENDED_FROM_YEAR
        and (args.to_year is None or y <= args.to_year)
        and _yaml_path(y).exists()
    ]

    table = Table(title=f"growth_factors from {SOURCE_LABEL} + {MHCLG_CT_LABEL}")
    table.add_column("FY")
    table.add_column("field")
    table.add_column("old", justify="right")
    table.add_column("new", justify="right")
    table.add_column("Δ", justify="right")

    written = 0
    for year in years:
        old = _read_existing(year)
        fields: dict[str, float] = {}
        for name, (_, legacy_trio) in _FIELDS.items():
            if name == "council_tax_growth":
                if year in council_tax:
                    fields[name] = council_tax[year]
            elif legacy_trio and year < args.from_year and name in old:
                # Outturn vintage already in the YAML takes precedence.
                fields[name] = old[name]
            elif name in rates.get(year, {}):
                fields[name] = rates[year][name]
        for name, nv in fields.items():
            ov = old.get(name)
            if ov is not None and abs(nv - ov) < 5e-6:
                continue  # unchanged — keep the diff table readable
            delta = "" if ov is None else f"{(nv - ov) * 100:+.3f}pp"
            table.add_row(
                f"{year}/{(year + 1) % 100:02d}",
                name,
                "—" if ov is None else f"{ov * 100:.3f}%",
                f"{nv * 100:.3f}%",
                delta,
            )
        if not args.dry_run:
            if _rewrite_block(year, _render_block(year, fields)):
                written += 1

    console.print(table)
    if args.dry_run:
        console.print(
            "[yellow]dry run — no files written. Unchanged values omitted.[/yellow]"
        )
    else:
        console.print(
            f"[green]wrote growth_factors to {written} YAML files ({years[0]}–{years[-1]}).[/green]"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
