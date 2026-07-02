"""Build the policy costings xlsx database from extracted.json."""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent

NON_NUMERIC = {"neg", "*", "-", "–", "—", "n/a", ""}


def to_number(raw):
    """Parse '£m' cell to float, or None if non-numeric (neg, *, etc.)."""
    s = raw.strip().lower().replace("м", "m")
    if s in NON_NUMERIC:
        return None
    s = s.replace("m", "").replace(",", "").replace("+", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def main():
    db = json.loads((HERE / "extracted.json").read_text())

    years = sorted(
        {y for ev in db for m in ev["measures"] for i in m["impacts"] for y in i["values"]}
    )

    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")

    def style_header(ws):
        for c in ws[1]:
            c.fill = head_fill
            c.font = head_font
        ws.freeze_panes = "A2"

    # README
    ws = wb.active
    ws.title = "README"
    readme = [
        ["UK fiscal event policy costings database, Budget 2016 to Budget 2025"],
        [""],
        ["Compiled from HM Treasury 'policy costings' documents published alongside each"],
        ["fiscal event. Each document sets out, per measure, the exchequer impact (£m) over"],
        ["the forecast horizon as certified by the OBR, on a National Accounts basis."],
        [""],
        ["Sheets:"],
        ["  events — one row per fiscal event, with source PDF URL"],
        ["  measures — one row per measure x impact-table row, fiscal years as columns,"],
        ["    with full text of the measure description, tax/cost base, costing methodology"],
        ["    (including static/behavioural detail for expanded-format measures), any prose"],
        ["    accompanying the impact table, and areas of uncertainty"],
        ["  long — tidy format: one row per measure x table row x year"],
        [""],
        ["Conventions:"],
        ["  Values are £ million. Negative = exchequer cost, positive = exchequer yield."],
        ["  'neg' / '*' = negligible (below the publication threshold); kept as text."],
        ["  Most measures have a single 'Exchequer impact' row; a few report component"],
        ["  rows (e.g. Budget 2025 business rates measures) or 'o/w' breakdowns."],
        [""],
        ["Classification (data/costings/classify.py):"],
        ["  policy_area — tax/spending head the measure changes (rule-based on titles)."],
        ["  in_model — whether the lever is represented in policyengine-uk-compiled."],
        ["    Labelled manually per measure against the model's parameters/variables:"],
        ["    yes = the specific lever is a model parameter (e.g. UC taper, IT thresholds);"],
        ["    partial = instrument modelled but lever/population not separately"],
        ["    representable (e.g. PIP eligibility criteria, draught relief);"],
        ["    no = outside the model (e.g. corporation tax, business rates, VAT scope)."],
        ["  incidence — statutory (who directly pays/receives): households, firms,"],
        ["    mixed, or public sector & other. This is not economic incidence."],
        [""],
        ["Coverage notes:"],
        ["  Every HMT policy costings document from Budget 2016 to Budget 2025 (17 events)."],
        ["  No costings document exists for: Spring Statements 2018/2019 (no measures),"],
        ["  Plan for Jobs (July 2020), Growth Plan (September 2022, no OBR forecast),"],
        ["  Spending Review 2025 (spending-only). Spending Review 2020 is included"],
        ["  ('Policy costings: November 2020')."],
        [""],
        ["  Measures below the negligible threshold in every year are excluded from the"],
        ["  source documents, so this is not an exhaustive list of all announcements."],
    ]
    for row in readme:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 95

    # events
    ws = wb.create_sheet("events")
    ws.append(["event_key", "event_name", "event_date", "n_measures", "source_pdf"])
    for ev in db:
        ws.append([ev["key"], ev["name"], ev["date"], len(ev["measures"]), ev["url"]])
    style_header(ws)
    for col, w in zip("ABCDE", [24, 38, 12, 12, 100]):
        ws.column_dimensions[col].width = w

    # measures (wide)
    ws = wb.create_sheet("measures")
    text_cols = [
        ("measure_description", "description"),
        ("tax_cost_base", "base"),
        ("costing_methodology", "costing"),
        ("impact_table_notes", "impact_notes"),
        ("areas_of_uncertainty", "uncertainty"),
    ]
    meta_cols = ["event_name", "event_date", "measure", "row_label",
                 "policy_area", "in_model", "incidence", "page"]
    n_meta = len(meta_cols)
    header = meta_cols + years + [name for name, _ in text_cols]
    ws.append(header)
    for ev in db:
        for m in ev["measures"]:
            rows = m["impacts"] or [{"label": "(no table found)", "values": {}}]
            for i in rows:
                row = [ev["name"], ev["date"], m["title"], i["label"],
                       m.get("policy_area", ""), m.get("in_model", ""),
                       m.get("incidence", ""), m["page_start"]]
                for y in years:
                    raw = i["values"].get(y, "")
                    num = to_number(raw)
                    row.append(num if num is not None else (raw or None))
                row += [m.get(key, "") for _, key in text_cols]
                ws.append(row)
    style_header(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 18
    for idx in range(n_meta + 1, n_meta + 1 + len(years)):
        ws.column_dimensions[get_column_letter(idx)].width = 10
    for k in range(len(text_cols)):
        ws.column_dimensions[get_column_letter(n_meta + 1 + len(years) + k)].width = 60
    for r in ws.iter_rows(min_row=2):
        r[2].alignment = wrap
        for c in r[n_meta + len(years) :]:
            c.alignment = wrap
        for c in r[n_meta : n_meta + len(years)]:
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0"

    # long
    ws = wb.create_sheet("long")
    ws.append(
        ["event_key", "event_name", "event_date", "measure", "row_label",
         "fiscal_year", "value_raw", "value_gbp_m"]
    )
    for ev in db:
        for m in ev["measures"]:
            for i in m["impacts"]:
                for y in sorted(i["values"]):
                    raw = i["values"][y]
                    if not raw:
                        continue
                    ws.append(
                        [ev["key"], ev["name"], ev["date"], m["title"], i["label"],
                         y, raw, to_number(raw)]
                    )
    style_header(ws)
    for col, w in zip("ABCDEFGH", [22, 30, 12, 60, 18, 12, 12, 12]):
        ws.column_dimensions[col].width = w
    for r in ws.iter_rows(min_row=2):
        if isinstance(r[7].value, (int, float)):
            r[7].number_format = "#,##0"

    out = HERE / "uk_policy_costings_2016_2025.xlsx"
    wb.save(out)

    n_rows = sum(len(m["impacts"]) for ev in db for m in ev["measures"])
    n_meas = sum(len(ev["measures"]) for ev in db)
    print(f"Wrote {out}: {len(db)} events, {n_meas} measures, {n_rows} impact rows, years {years[0]}..{years[-1]}")

    # compact JSON for the docs site explorer
    site_db = []
    for ev in db:
        site_db.append({
            "key": ev["key"], "name": ev["name"], "date": ev["date"], "url": ev["url"],
            "measures": [
                {
                    "title": m["title"],
                    "page": m["page_start"],
                    "area": m.get("policy_area", ""),
                    "inModel": m.get("in_model", ""),
                    "incidence": m.get("incidence", ""),
                    "impacts": [
                        {"label": i["label"],
                         "values": {y: {"raw": v, "num": to_number(v)} for y, v in i["values"].items() if v}}
                        for i in m["impacts"]
                    ],
                    "description": m.get("description", ""),
                    "base": m.get("base", ""),
                    "costing": m.get("costing", ""),
                    "impactNotes": m.get("impact_notes", ""),
                    "uncertainty": m.get("uncertainty", ""),
                }
                for m in ev["measures"]
            ],
        })
    site_out = HERE.parent.parent / "site" / "public" / "costings.json"
    site_out.write_text(json.dumps(site_db, separators=(",", ":")))
    print(f"Wrote {site_out} ({site_out.stat().st_size/1e6:.1f} MB)")


if __name__ == "__main__":
    main()
