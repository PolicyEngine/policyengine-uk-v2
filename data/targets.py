"""Fetch OBR EFO calibration targets and write data/targets/YYYY_YY.yaml files.

Downloads the OBR Economic and Fiscal Outlook detailed forecast tables
(receipts + expenditure xlsx) and extracts annual £bn totals for each program
in the calibration audit (data/calibrate.py). One YAML is written per fiscal
year in the forecast horizon.

Update EFO_EDITION when a new EFO is published.

Usage:
    python data/targets.py               # write all years in the forecast
    python data/targets.py --year 2025   # single year
"""

from __future__ import annotations

import argparse
import re
import urllib.request
from pathlib import Path

import openpyxl
import yaml
from rich.console import Console
from rich.table import Table

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = REPO_ROOT / "data" / "raw" / "obr"
TARGETS_DIR = REPO_ROOT / "data" / "targets"

EFO_EDITION = "march-2026"
# DWP benefit expenditure and caseload tables, consistent with the same EFO.
DWP_URL = (
    "https://assets.publishing.service.gov.uk/media/69dcdc8c6b695d635c34dcc4/"
    "outturn-and-forecast-tables-spring-forecast-2026.xlsx"
)
FILES = {
    "receipts": f"https://obr.uk/download/{EFO_EDITION}-economic-and-fiscal-outlook-detailed-forecast-tables-receipts/",
    "expenditure": f"https://obr.uk/download/{EFO_EDITION}-economic-and-fiscal-outlook-detailed-forecast-tables-expenditure/",
    "dwp": DWP_URL,
}
# OBR tables are £bn, DWP tables £m.
SCALE = {"receipts": 1.0, "expenditure": 1.0, "dwp": 1e-3}

# program → (file, sheet, row label prefix, sum all matching rows)
# Benefits split across the welfare cap (e.g. UC, housing benefit) appear as
# two rows with the same label, so sum_all combines them.
SPEC: dict[str, tuple[str, str, str, bool]] = {
    "income_tax": ("receipts", "3.8", "Income tax (gross of tax credits)", False),
    "employee_ni": ("receipts", "3.4", "Class 1 Employee NICs", False),
    "employer_ni": ("receipts", "3.4", "Class 1 Employer NICs", False),
    "vat": ("receipts", "3.8", "Value added tax", False),
    "capital_gains_tax": ("receipts", "3.8", "Capital gains tax", False),
    "stamp_duty": ("receipts", "3.8", "Stamp duty land tax", False),
    "council_tax": ("receipts", "3.8", "Council tax", False),
    "universal_credit": ("expenditure", "4.9", "Universal credit", True),
    "child_benefit": ("expenditure", "4.9", "Child benefit", False),
    "state_pension": ("expenditure", "4.9", "State pension", False),
    "pension_credit": ("expenditure", "4.9", "Pension credit", False),
    "housing_benefit": ("expenditure", "4.9", "Housing benefit", True),
    "carers_allowance": ("expenditure", "4.9", "Carer's allowance", False),
    # Legacy benefits from DWP tables (OBR doesn't separate these). Each sheet
    # repeats labels in real-terms and caseload blocks below the nominal one,
    # so first match (nominal £m) wins. CTC/WTC only exist combined; the
    # tax_credits key is summed from both engine fields in calibrate.py.
    "income_support": ("dwp", "Income Support", "Total", False),
    "esa_income_related": ("dwp", "Incapacity benefits", "Employment and Support Allowance (income based)", False),
    "jsa_income_based": ("dwp", "Unemployment benefits", "of which income-based", False),
    "tax_credits": ("dwp", "Non-DWP Welfare", "of which Child Tax Credit and Working Tax Credit", False),
}

console = Console()


def download(name: str) -> Path:
    path = RAW_DIR / f"{name}.xlsx"
    if path.exists():
        console.print(f"  [dim]{name} cached at {path}[/dim]")
        return path
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    console.print(f"  downloading {FILES[name]}")
    req = urllib.request.Request(FILES[name], headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        path.write_bytes(r.read())
    return path


def _year_columns(ws) -> dict[int, int]:
    """Find the header row of fiscal year labels ('2025-26') → {year: column}."""
    for row in range(1, 10):
        cols = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and re.fullmatch(r"\d{4}[-/]\d{2}", val.strip()):
                cols[int(val.strip()[:4])] = col
        if cols:
            return cols
    raise SystemExit(f"No fiscal year header row found in sheet {ws.title}")


def _read_program(ws, label: str, sum_all: bool, year_cols: dict[int, int]) -> dict[int, float]:
    values: dict[int, float] = {}
    matched = 0
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=2).value
        if not (isinstance(cell, str) and cell.strip().startswith(label)):
            continue
        matched += 1
        for year, col in year_cols.items():
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                values[year] = values.get(year, 0.0) + float(val)
        if not sum_all:
            break
    if matched == 0:
        raise SystemExit(f"Row '{label}' not found in sheet {ws.title} — has the EFO layout changed?")
    return values


def fetch_targets() -> dict[int, dict[str, float]]:
    """Return {year: {program: £bn}} parsed from the OBR tables."""
    books = {name: openpyxl.load_workbook(download(name), data_only=True) for name in FILES}
    targets: dict[int, dict[str, float]] = {}
    for program, (file, sheet, label, sum_all) in SPEC.items():
        ws = books[file][sheet]
        for year, value in _read_program(ws, label, sum_all, _year_columns(ws)).items():
            targets.setdefault(year, {})[program] = round(value * SCALE[file], 3)
    for wb in books.values():
        wb.close()
    # DWP tables go back decades; only keep years every program covers.
    return {y: progs for y, progs in targets.items() if len(progs) == len(SPEC)}


def write_yaml(year: int, programs: dict[str, float]) -> Path:
    TARGETS_DIR.mkdir(parents=True, exist_ok=True)
    path = TARGETS_DIR / f"{year}_{str(year + 1)[-2:]}.yaml"
    header = (
        f"# Calibration targets for {year}/{str(year + 1)[-2:]} (£bn).\n"
        f"# Sources: OBR EFO {EFO_EDITION} detailed forecast tables (receipts 3.4/3.8,\n"
        f"# expenditure 4.9); DWP benefit expenditure and caseload tables (legacy benefits).\n"
        f"# Generated by data/targets.py — do not edit by hand.\n"
    )
    path.write_text(header + yaml.safe_dump(programs, sort_keys=True))
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int, help="Only write this fiscal year")
    args = parser.parse_args()

    targets = fetch_targets()
    years = [args.year] if args.year else sorted(targets)
    if args.year and args.year not in targets:
        raise SystemExit(f"Year {args.year} not in EFO horizon ({min(targets)}–{max(targets)})")

    table = Table(title=f"OBR EFO {EFO_EDITION} targets (£bn)", show_header=True)
    table.add_column("Program", style="bold")
    for y in years:
        table.add_column(f"{y}-{str(y + 1)[-2:]}", justify="right")
    for program in SPEC:
        table.add_row(program.replace("_", " "), *[f"{targets[y][program]:.1f}" for y in years])
    console.print(table)

    for y in years:
        path = write_yaml(y, targets[y])
        console.print(f"  wrote {path.relative_to(REPO_ROOT)}")

    console.print("[green]Done.[/green]")


if __name__ == "__main__":
    main()
