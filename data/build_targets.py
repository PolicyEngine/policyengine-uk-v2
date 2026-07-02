"""Build calibration targets JSON covering 2010–2024.

Sources:
  - HMRC receipts ODS (data/raw/hmrc_receipts.ods): income tax, NI, CGT, VAT, stamp duty
  - DWP/OBR expenditure (data/raw/obr/dwp.xlsx): all major benefits
  - Eurostat nama_10_co3_p3 + ONS QNA: COICOP household consumption at current prices
  - DWP/OBR expenditure (data/raw/obr/dwp.xlsx): benefit claimant caseloads

Usage:
    python data/build_targets.py
    python data/build_targets.py --years 2020 2023
    python data/build_targets.py --no-download   # skip re-downloading raw files
"""

from __future__ import annotations

import argparse
import json
import os
import re
import urllib.request
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from uprate import cumulative_factor

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = REPO_ROOT / "data" / "raw"
OUT_PATH = REPO_ROOT / "data" / "calibration_targets.json"

HMRC_ODS_URL = (
    "https://assets.publishing.service.gov.uk/media/"
    "6a0c6b6efcae986635db91f0/NS_Table.ods"
)
ONS_QNA_URL = (
    "https://www.ons.gov.uk/file?uri=/economy/grossdomesticproductgdp/datasets/"
    "uksecondestimateofgdpdatatables/quarter1jantomar2026firstestimate/"
    "firstquarterlyestimatedatatablescorrected.xlsx"
)

# EFRS starts at 2016 (data/efrs.py YEARS): pre-2013 clean FRS frames carry no
# housing benefit, so years whose FRS pool reaches before 2013 are dropped. 2015
# (pool back to FRS 2013) is also dropped as the thinnest real year.
TARGET_YEARS = list(range(2016, 2025))

# Forecast horizon: the last year with real source data, and the OBR EFO years we
# project onto by uprating the latest real targets (data/uprate.py indices).
LATEST_REAL_YEAR = 2024
FORECAST_YEARS = list(range(2025, 2030))


# ── helpers ─────────────────────────────────────────────────────────────────


def _download(url: str, dest: Path) -> None:
    if dest.exists():
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        dest.write_bytes(r.read())


def _ods_rows(path: Path, sheet_name: str) -> list[list]:
    import odf.opendocument
    import odf.table

    def cell_text(cell):
        parts = []
        for el in cell.childNodes:
            if hasattr(el, "data"):
                parts.append(el.data)
            elif el.qname[1] == "p":
                for sub in el.childNodes:
                    if hasattr(sub, "data"):
                        parts.append(sub.data)
        return "".join(parts)

    doc = odf.opendocument.load(str(path))
    for s in doc.spreadsheet.getElementsByType(odf.table.Table):
        if s.getAttribute("name") == sheet_name:
            rows = []
            for row in s.getElementsByType(odf.table.TableRow):
                cells = []
                for cell in row.getElementsByType(odf.table.TableCell):
                    repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                    cells.extend([cell_text(cell)] * repeat)
                rows.append(cells)
            return rows
    raise KeyError(sheet_name)


def _xls_col(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    return v


# ── HMRC receipts ────────────────────────────────────────────────────────────


def load_hmrc_receipts(path: Path) -> dict[int, dict[str, float]]:
    """Return {fiscal_year_start: {field: £bn}}."""
    rows = _ods_rows(path, "Receipts_Annually")

    # Row index 5 (0-based) is the header
    header = rows[5]
    col = {
        "income_tax": next(i for i, h in enumerate(header) if h == "Income Tax"),
        "employee_ni": next(i for i, h in enumerate(header) if "EMP'ee" in h),
        "employer_ni": next(i for i, h in enumerate(header) if "EMP'er" in h),
        "total_nic": next(
            i for i, h in enumerate(header) if h == "National Insurance Contributions"
        ),
        "capital_gains_tax": next(
            i for i, h in enumerate(header) if h == "Capital Gains Tax"
        ),
        "vat": next(i for i, h in enumerate(header) if h == "Value Added Tax"),
        "stamp_duty": next(
            i for i, h in enumerate(header) if h == "Stamp Duty Land Tax"
        ),
    }

    out: dict[int, dict[str, float]] = {}
    for row in rows[6:]:
        m = re.match(r"(\d{4}) to (\d{4})", row[0]) if row else None
        if not m:
            continue
        fy = int(m.group(1))
        if fy < 2010 or fy > 2024:
            continue

        def get(c: int) -> float | None:
            v = row[c].replace(",", "").replace(" ", "").strip()
            try:
                return float(v) / 1000.0  # £m -> £bn
            except ValueError:
                return None

        # NI: use employer+employee if available, else split total proportionally
        emp_ee = get(col["employee_ni"])
        emp_er = get(col["employer_ni"])
        total_nic = get(col["total_nic"])

        if emp_ee is not None and emp_er is not None:
            employee_ni = emp_ee
            employer_ni = emp_er
        elif total_nic is not None:
            # 2010-2014: split not available; use 2015 ratio of employee/(employee+employer)
            # 2015 ratio from the data: emp_ee=44.792, emp_er=61.969, total=~107
            # employee share ≈ 0.420
            employee_ni = total_nic * 0.420
            employer_ni = total_nic * 0.580
        else:
            employee_ni = employer_ni = None

        out[fy] = {
            "income_tax": get(col["income_tax"]),
            "employee_ni": employee_ni,
            "employer_ni": employer_ni,
            "capital_gains_tax": get(col["capital_gains_tax"]),
            "vat": get(col["vat"]),
            "stamp_duty": get(col["stamp_duty"]),
        }

    return out


def load_obr_databank_receipts(path: Path) -> dict[int, dict[str, float]]:
    """Tax receipts (£bn) from the OBR public finances databank 'Receipts (£bn)' sheet.

    Covers 1999-00 onwards — used to backfill years the HMRC monthly bulletin
    (which starts 2005-06) lacks. NICs are a single total here, so split with the
    same 0.420 employee share the HMRC loader uses for years without a breakdown.
    Values agree with HMRC to within a few per cent at the overlap.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["Receipts (£bn)"]

    # Column layout fixed by the databank header (row 4):
    cols = {
        "vat": 3,
        "stamp_duty": 6,
        "cgt": 20,
        "nics": 28,
        "paye": 17,
        "sa": 18,
        "other_it": 19,
    }

    out: dict[int, dict[str, float]] = {}
    for r in range(7, ws.max_row + 1):
        y = ws.cell(row=r, column=2).value
        if not (isinstance(y, str) and re.match(r"\d{4}-\d{2}", y)):
            continue
        fy = int(y[:4])

        def g(key: str) -> float:
            v = ws.cell(row=r, column=cols[key]).value
            return float(v) if isinstance(v, (int, float)) else 0.0

        nics = g("nics")
        out[fy] = {
            "income_tax": g("paye") + g("sa") + g("other_it"),
            "employee_ni": nics * 0.420,
            "employer_ni": nics * 0.580,
            "capital_gains_tax": g("cgt"),
            "vat": g("vat"),  # net of refunds, matching HMRC's VAT line
            "stamp_duty": g("stamp_duty"),
        }
    return out


def load_efo_labour_market(path: Path) -> dict[int, dict[str, float]]:
    """Employment and ILO unemployment levels (counts) by fiscal year from the
    OBR EFO economy table 1.6.

    The table is quarterly (rows "YYYYQn"); column 2 is employment (16+, millions)
    and column 5 is ILO unemployment (16+, millions). Each fiscal year is the mean
    of Q2..Q1 (e.g. 2010 = 2010Q2..2011Q1), matching the FRS year convention. Levels
    returned as raw person counts (millions x 1e6).
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["1.6"]

    quarters: dict[tuple[int, int], dict[str, float]] = {}
    for row in ws.iter_rows(values_only=True):
        label = row[1]
        m = re.match(r"(\d{4})Q([1-4])", str(label)) if label else None
        if not m:
            continue
        emp, unemp = row[2], row[5]
        if isinstance(emp, (int, float)) and isinstance(unemp, (int, float)):
            quarters[(int(m.group(1)), int(m.group(2)))] = {
                "employed": float(emp) * 1e6,
                "unemployed": float(unemp) * 1e6,
            }

    out: dict[int, dict[str, float]] = {}
    for fy in TARGET_YEARS:
        fy_quarters = [(fy, 2), (fy, 3), (fy, 4), (fy + 1, 1)]
        vals = [quarters[q] for q in fy_quarters if q in quarters]
        if len(vals) == 4:
            out[fy] = {
                "employed": sum(v["employed"] for v in vals) / 4.0,
                "unemployed": sum(v["unemployed"] for v in vals) / 4.0,
            }
    return out


# ── SPI income distribution (Table 3.6) ─────────────────────────────────────

# HMRC SPI Table 3.6 by total-income band, per income source: number of
# individuals (thousands) and amount (£m). Two column layouts across the file
# vintages: xls/xlsx (wide, with spacer columns) and ods (compact). Each source
# maps to an engine person variable.
_SPI_SOURCES = [  # (key, person_variable)
    ("self_employment", "self_employment_income"),
    ("employment", "employment_income"),
    ("state_pension", "state_pension"),
    ("private_pension", "private_pension_income"),
]
_SPI_COLS_WIDE = {
    "band": 0,
    "self_employment": (2, 3),
    "employment": (6, 7),
    "state_pension": (10, 11),
    "private_pension": (14, 15),
}
_SPI_COLS_ODS = {
    "band": 0,
    "self_employment": (1, 2),
    "employment": (4, 5),
    "state_pension": (7, 8),
    "private_pension": (10, 11),
}


def load_spi_table_3_6(path: Path) -> list[dict]:
    """Return banded distribution rows from one SPI Table 3.6 file.

    Each row: {lo, hi, <src>_n (count), <src>_a (£m amount)} for each source in
    _SPI_SOURCES, where the band runs [lo, hi) over total income. The top band's
    hi is +inf. Rows whose band cell is not a positive number are skipped.
    """
    if path.suffix == ".xls":
        engine, cols, sheet = "xlrd", _SPI_COLS_WIDE, 0
    elif path.suffix == ".xlsx":
        engine, cols, sheet = "openpyxl", _SPI_COLS_WIDE, 0
    else:
        xl = pd.ExcelFile(path, engine="odf")
        sheet = next(s for s in ("Table_3_6", "T3_6") if s in xl.sheet_names)
        engine, cols = "odf", _SPI_COLS_ODS

    df = pd.read_excel(path, sheet_name=sheet, header=None, engine=engine)

    def num(v: object) -> float | None:
        return float(v) if isinstance(v, (int, float)) and pd.notna(v) else None

    rows: list[dict] = []
    for i in range(len(df)):
        lo = num(df.iloc[i, cols["band"]])
        if lo is None or lo <= 0:
            continue
        rec: dict[str, float] = {"lo": lo}
        for key, _ in _SPI_SOURCES:
            cc, ca = cols[key]
            rec[f"{key}_n"] = num(df.iloc[i, cc]) or 0.0
            rec[f"{key}_a"] = num(df.iloc[i, ca]) or 0.0
        rows.append(rec)
    # The very first numeric row in the xls/xlsx vintages is a stray header echo;
    # drop any leading row with no source counts at all.
    rows = [r for r in rows if any(r[f"{k}_n"] > 0 for k, _ in _SPI_SOURCES)]
    for k in range(len(rows)):
        rows[k]["hi"] = rows[k + 1]["lo"] if k + 1 < len(rows) else float("inf")
    return rows


def load_spi_distributions(spi_dir: Path) -> dict[int, list[dict]]:
    """Load all SPI Table 3.6 files keyed by fiscal-year start (2010 = 2010-11)."""
    out: dict[int, list[dict]] = {}
    for path in sorted(spi_dir.glob("Table_3_6_*")):
        m = re.search(r"_(\d{4})\.", path.name)
        if m:
            out[int(m.group(1))] = load_spi_table_3_6(path)
    return out


# Collapse SPI Tables 3.6/3.7 bands at/above this total-income threshold into one.
# The public tape is disclosure-controlled in the top bands: FACT scaling inflates
# the grossed counts so the per-band mean income the survey carries no longer
# matches the published count/amount pair (self-employment £523k+/£1m+ sit ~14%
# below their target mean). Reweighting can't fix a per-record mean, so the over-
# tight sub-band constraints are unfittable. Pooling them into one band leaves a
# single aggregate count+amount the survey can satisfy by reweighting toward its
# highest earners. The floor is a raw SPI band edge (compared before earnings-
# uprating); the raw bands run …£200k/£300k/£500k/£1m. It sits at £300k (one band
# below the disclosure-control onset at £500k): the pooled £500k+£1m band alone is
# still too thin to fit count and amount together under the 10x clamp, so the
# well-represented £300k band is folded in to give the count constraint slack. The
# pension count-drop below is coupled to this floor, so the inflated top-band
# pension headcounts stay dropped.
_SPI_TOP_POOL_FLOOR = 300_000.0


def _pool_spi_top_rows(rows: list[dict], sources: list[tuple[str, str]]) -> list[dict]:
    """Collapse all bands with lo >= _SPI_TOP_POOL_FLOOR into one combined band,
    summing each source's count and amount; bands below the floor pass through."""
    low = [r for r in rows if r["lo"] < _SPI_TOP_POOL_FLOOR]
    high = [r for r in rows if r["lo"] >= _SPI_TOP_POOL_FLOOR]
    if len(high) <= 1:
        return rows
    merged = {"lo": min(r["lo"] for r in high), "hi": max(r["hi"] for r in high)}
    for key, _ in sources:
        merged[f"{key}_n"] = sum(r[f"{key}_n"] for r in high)
        merged[f"{key}_a"] = sum(r[f"{key}_a"] for r in high)
    return [*low, merged]


# ── SPI investment income (Table 3.7) ───────────────────────────────────────

# HMRC SPI Table 3.7 by total-income band: number of individuals (thousands) and
# amount (£m) of property, interest and dividend income — the unearned-income
# sources Table 3.6 omits. Two layouts: compact ods (number/amount in adjacent
# columns) and wide xlsx (a spacer column splits each source's number/amount).
# "Other income" is deliberately skipped: it has no clean engine counterpart.
_SPI37_SOURCES = [  # (key, person_variable)
    ("property", "property_income"),
    ("interest", "savings_interest"),
    ("dividends", "dividend_income"),
]
_SPI37_COLS_WIDE = {
    "band": 0,
    "property": (2, 3),
    "interest": (6, 7),
    "dividends": (10, 11),
}
_SPI37_COLS_ODS = {
    "band": 0,
    "property": (1, 2),
    "interest": (4, 5),
    "dividends": (7, 8),
}


def load_spi_table_3_7(path: Path) -> list[dict]:
    """Return banded investment-income rows from one SPI Table 3.7 file.

    Each row: {lo, hi, <src>_n (count, thousands), <src>_a (£m)} for each source
    in _SPI37_SOURCES, band [lo, hi) over total income; the top band's hi is +inf.
    """
    if path.suffix in (".xlsx", ".xls"):
        engine = "openpyxl" if path.suffix == ".xlsx" else "xlrd"
        cols = _SPI37_COLS_WIDE
        names = pd.ExcelFile(path, engine=engine).sheet_names
        sheet = next((s for s in names if "3.7" in s or "3_7" in s), names[0])
    else:
        xl = pd.ExcelFile(path, engine="odf")
        sheet = next(s for s in ("Table_3_7", "T3_7") if s in xl.sheet_names)
        engine, cols = "odf", _SPI37_COLS_ODS

    df = pd.read_excel(path, sheet_name=sheet, header=None, engine=engine)

    def num(v: object) -> float | None:
        return float(v) if isinstance(v, (int, float)) and pd.notna(v) else None

    rows: list[dict] = []
    for i in range(len(df)):
        lo = num(df.iloc[i, cols["band"]])
        if lo is None or lo <= 0:
            continue
        rec: dict[str, float] = {"lo": lo}
        for key, _ in _SPI37_SOURCES:
            cc, ca = cols[key]
            rec[f"{key}_n"] = num(df.iloc[i, cc]) or 0.0
            rec[f"{key}_a"] = num(df.iloc[i, ca]) or 0.0
        rows.append(rec)
    rows = [r for r in rows if any(r[f"{k}_n"] > 0 for k, _ in _SPI37_SOURCES)]
    for k in range(len(rows)):
        rows[k]["hi"] = rows[k + 1]["lo"] if k + 1 < len(rows) else float("inf")
    return rows


def load_spi_investment(spi_dir: Path) -> dict[int, list[dict]]:
    """Load all SPI Table 3.7 files keyed by fiscal-year start (2016 = 2016-17)."""
    out: dict[int, list[dict]] = {}
    for path in sorted(spi_dir.glob("Table_3_7_*")):
        m = re.search(r"_(\d{4})\.", path.name)
        if m:
            out[int(m.group(1))] = load_spi_table_3_7(path)
    return out


# Total-income floor below which the SPI 3.7 bands are pooled into one combined
# band per source. Each sub-floor band holds only a few hundred £m of investment
# income spread over thin survey cells, so the engine's small absolute
# misallocation there becomes a huge relative error (dividends in the personal-
# allowance band ran +170% to +3700%) that the magnitude-blind RMSRE squares.
# Pooling below £30k lifts the denominator: combined dividend/property errors
# fall to single digits while interest (already fitting) is unaffected.
_SPI37_POOL_FLOOR = 30000.0


def _pool_spi37_rows(rows: list[dict]) -> list[dict]:
    """Collapse all bands with lo < _SPI37_POOL_FLOOR into one combined band,
    summing each source's count and amount; bands above the floor pass through."""
    low = [r for r in rows if r["lo"] < _SPI37_POOL_FLOOR]
    high = [r for r in rows if r["lo"] >= _SPI37_POOL_FLOOR]
    if len(low) <= 1:
        return rows
    merged = {"lo": min(r["lo"] for r in low), "hi": max(r["hi"] for r in low)}
    for key, _ in _SPI37_SOURCES:
        merged[f"{key}_n"] = sum(r[f"{key}_n"] for r in low)
        merged[f"{key}_a"] = sum(r[f"{key}_a"] for r in low)
    return [merged, *high]


def build_spi_investment_targets(
    spi37: dict[int, list[dict]], earnings_index: dict[int, float], years: list[int]
) -> list[dict]:
    """Banded SPI investment-income targets (Table 3.7): per band per source, a
    count (count_nonzero) and an amount (sum), filtered on baseline_total_income.

    Mirrors build_spi_targets: years beyond the latest vintage reuse its bands
    with counts held and amounts grown by the EFO average-earnings ratio. Bands
    below _SPI37_POOL_FLOOR are pooled (see _pool_spi37_rows).
    """
    label = "HMRC SPI Table 3.7 (investment income distribution)"
    avail = sorted(spi37)
    out: list[dict] = []
    for yr in years:
        if yr in spi37:
            rows, amt_factor = spi37[yr], 1.0
        elif avail and yr > avail[-1]:
            base = avail[-1]
            rows = spi37[base]
            amt_factor = earnings_index[yr] / earnings_index[base]
        else:
            continue
        for r in _pool_spi_top_rows(_pool_spi37_rows(rows), _SPI37_SOURCES):
            # Beyond the latest vintage the band edges uprate with earnings too,
            # not just the amounts (see build_spi_targets).
            lo = r["lo"] * amt_factor
            hi = r["hi"] if r["hi"] == float("inf") else r["hi"] * amt_factor
            band = {
                "variable": "total_income",
                "min": round(lo, 0),
                "max": None if hi == float("inf") else round(hi, 0),
            }
            tag = int(round(lo))
            for key, variable in _SPI37_SOURCES:
                count, amount = r[f"{key}_n"], r[f"{key}_a"]
                if count > 0:
                    out.append(
                        {
                            "name": f"spi37_{key}_count_{tag}_{yr}",
                            "variable": variable,
                            "entity": "person",
                            "aggregation": "count_nonzero",
                            "filter": band,
                            "benunit_filter": None,
                            "value": round(count * 1000.0, 0),
                            "source": label,
                            "year": yr,
                            "holdout": False,
                        }
                    )
                if amount > 0:
                    out.append(
                        {
                            "name": f"spi37_{key}_amount_{tag}_{yr}",
                            "variable": variable,
                            "entity": "person",
                            "aggregation": "sum",
                            "filter": band,
                            "benunit_filter": None,
                            "value": round(amount * 1e6 * amt_factor, 0),
                            "source": label,
                            "year": yr,
                            "holdout": False,
                        }
                    )
    return out


def load_efo_earnings_index(path: Path) -> dict[int, float]:
    """Average-earnings index (2008Q1=100) by fiscal year from EFO table 1.6.

    Column 17 is the average weekly earnings index level; column 16 is its
    growth rate — using the rate here silently deflates extrapolated SPI bands.

    Used to project SPI income amounts into years the SPI does not yet cover
    (the latest SPI is 2023-24); 2024 amounts grow by index[2024]/index[2023].
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["1.6"]
    quarters: dict[tuple[int, int], float] = {}
    for row in ws.iter_rows(values_only=True):
        m = re.match(r"(\d{4})Q([1-4])", str(row[1])) if row[1] else None
        if m and len(row) > 17 and isinstance(row[17], (int, float)):
            quarters[(int(m.group(1)), int(m.group(2)))] = float(row[17])
    out: dict[int, float] = {}
    for fy in TARGET_YEARS:
        vals = [
            quarters[q]
            for q in [(fy, 2), (fy, 3), (fy, 4), (fy + 1, 1)]
            if q in quarters
        ]
        if len(vals) == 4:
            out[fy] = sum(vals) / 4.0
    return out


def build_spi_targets(
    spi: dict[int, list[dict]], earnings_index: dict[int, float], years: list[int]
) -> list[dict]:
    """Banded SPI income-distribution targets: per band per source, a count
    (count_nonzero of the source variable) and an amount (sum). Each target
    carries a person-level band filter on baseline_total_income [lo, hi).

    Years beyond the latest SPI vintage reuse that year's bands with counts held
    and amounts grown by the EFO average-earnings ratio.
    """
    label = "HMRC SPI Table 3.6 (income distribution)"
    avail = sorted(spi)
    out: list[dict] = []
    for yr in years:
        if yr in spi:
            rows, amt_factor = spi[yr], 1.0
        elif avail and yr > avail[-1]:
            base = avail[-1]
            rows = spi[base]
            amt_factor = earnings_index[yr] / earnings_index[base]
        else:
            continue
        for r in _pool_spi_top_rows(rows, _SPI_SOURCES):
            # Beyond the latest vintage the band edges uprate with earnings too,
            # not just the amounts — otherwise the £-thresholds stay frozen while
            # the amounts grow, mis-aligning the band a record's income falls in.
            lo = r["lo"] * amt_factor
            hi = r["hi"] if r["hi"] == float("inf") else r["hi"] * amt_factor
            band = {
                "variable": "total_income",
                "min": round(lo, 0),
                "max": None if hi == float("inf") else round(hi, 0),
            }
            tag = int(round(lo))
            for key, variable in _SPI_SOURCES:
                count, amount = r[f"{key}_n"], r[f"{key}_a"]
                # The public SPI tape is disclosure-controlled in the top bands:
                # FACT jumps ~10→~32 per record, so its grossed pension headcount
                # (~20k in the £1m+ band) is ~10× the published count (~2k) while
                # the amount agrees. Injected records carry the tape's many-small-
                # pensions shape, so a count target there is unreachable — keep only
                # the amount for pension bands at/above the top-pool floor (£314k).
                if (
                    key in ("state_pension", "private_pension")
                    and r["lo"] >= _SPI_TOP_POOL_FLOOR
                ):
                    count = 0
                if count > 0:
                    out.append(
                        {
                            "name": f"spi_{key}_count_{tag}_{yr}",
                            "variable": variable,
                            "entity": "person",
                            "aggregation": "count_nonzero",
                            "filter": band,
                            "benunit_filter": None,
                            "value": round(count * 1000.0, 0),
                            "source": label,
                            "year": yr,
                            "holdout": False,
                        }
                    )
                if amount > 0:
                    out.append(
                        {
                            "name": f"spi_{key}_amount_{tag}_{yr}",
                            "variable": variable,
                            "entity": "person",
                            "aggregation": "sum",
                            "filter": band,
                            "benunit_filter": None,
                            "value": round(amount * 1e6 * amt_factor, 0),
                            "source": label,
                            "year": yr,
                            "holdout": False,
                        }
                    )
    return out


# ── DWP benefits ─────────────────────────────────────────────────────────────


def load_dwp_benefits(path: Path) -> dict[int, dict[str, float]]:
    """Return {fiscal_year_start: {field: £bn}} from Table 1a (£m nominal)."""
    wb = openpyxl.load_workbook(str(path), data_only=True)

    # Table 1a = nominal £m
    ws = wb["Table 1a"]
    row2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    yr_cols: dict[int, int] = {}
    for c, v in enumerate(row2, 1):
        if isinstance(v, str) and re.match(r"\d{4}/\d{2}", v.strip()):
            yr_cols[int(v.strip()[:4])] = c

    prog_rows = {
        "attendance_allowance": 7,
        "carers_allowance": 9,
        "disability_living_allowance": 18,
        "esa_income_related": 24,
        "housing_benefit": 29,
        "income_support": 35,
        "jsa_income_based": 43,
        "personal_independence_payment": 55,
        "pension_credit": 54,
        "state_pension": 63,
        "universal_credit": 72,
    }

    out: dict[int, dict[str, float]] = {}
    # Read every year column the table carries — the spring-forecast tables
    # include forecast years, which the UC £ total override consumes.
    for yr in sorted(yr_cols):
        row: dict[str, float] = {}
        for prog, row_idx in prog_rows.items():
            v = ws.cell(row=row_idx, column=yr_cols[yr]).value
            if isinstance(v, (int, float)):
                row[prog] = v / 1000.0  # £m -> £bn

        # Child benefit is in "UK welfare " sheet
        ws_uk = wb["UK welfare "]
        row2_uk = [
            ws_uk.cell(row=2, column=c).value for c in range(1, ws_uk.max_column + 1)
        ]
        yr_cols_uk: dict[int, int] = {}
        for c, v in enumerate(row2_uk, 1):
            if isinstance(v, str) and re.match(r"\d{4}/\d{2}", v.strip()):
                yr_cols_uk[int(v.strip()[:4])] = c

        if yr in yr_cols_uk:
            cb = ws_uk.cell(row=10, column=yr_cols_uk[yr]).value
            tc = ws_uk.cell(row=13, column=yr_cols_uk[yr]).value
            if isinstance(cb, (int, float)):
                row["child_benefit"] = cb
            if isinstance(tc, (int, float)):
                row["tax_credits"] = tc

        out[yr] = row

    return out


# ── COICOP consumption ───────────────────────────────────────────────────────


def load_coicop(ons_qna_path: Path) -> dict[int, dict[str, float]]:
    """Return {calendar_year: {field: £bn}} using Eurostat + ONS QNA."""

    # --- ONS QNA: total HHFCE at current prices (£m) and CVM breakdown ---
    wb = openpyxl.load_workbook(str(ons_qna_path), data_only=True)

    ws_c1 = wb["C1 EXPENDITURE"]
    c1: dict[int, float] = {}
    in_t1 = False
    for row in range(1, ws_c1.max_row + 1):
        v = ws_c1.cell(row=row, column=1).value
        if v == "Table 1: Annual":
            in_t1 = True
            continue
        if in_t1 and isinstance(v, int) and v > 1900:
            val = ws_c1.cell(row=row, column=2).value
            if isinstance(val, (int, float)):
                c1[v] = float(val)
        if in_t1 and v is None and len(c1) > 3:
            break

    ws_e3 = wb["E3 EXPENDITURE"]
    e3: dict[int, list] = {}
    in_t1 = False
    for row in range(1, ws_e3.max_row + 1):
        v = ws_e3.cell(row=row, column=1).value
        if v == "Table 1: Annual":
            in_t1 = True
            continue
        if in_t1 and isinstance(v, int) and v > 1900:
            vals = [ws_e3.cell(row=row, column=c).value for c in range(1, 17)]
            e3[v] = vals
        if in_t1 and v is None and len(e3) > 3:
            break

    # --- Eurostat: UK nominal COICOP £m 2010-2019 ---
    # The Eurostat API is occasionally unavailable; if the fetch or parse fails,
    # fall back to a no-op reader so pre-2020 years simply get no consumption
    # target (rather than aborting the whole target build). 2020+ years use the
    # CVM-share path below regardless.
    try:
        r = requests.get(
            "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/nama_10_co3_p3",
            params={"geo": "UK", "unit": "CP_MNAC", "freq": "A", "format": "JSON"},
            timeout=30,
        )
        r.raise_for_status()
        estat = r.json()
        coicop_idx = estat["dimension"]["coicop"]["category"]["index"]
        time_idx = estat["dimension"]["time"]["category"]["index"]
        n_time = estat["size"][4]
        estat_vals = estat["value"]

        def estat_get(code: str, year: int) -> float | None:
            cp = coicop_idx.get(code)
            tp = time_idx.get(str(year))
            if cp is None or tp is None:
                return None
            return estat_vals.get(str(cp * n_time + tp))
    except (requests.RequestException, ValueError, KeyError):

        def estat_get(code: str, year: int) -> float | None:
            return None

    # col indices in E3 row (0-based from row list): 0=yr,1=national,2=net_tourism,3=domestic,
    # 4=food,5=alc+tob,6=clothing,7=housing,8=furnishings,9=health,10=transport,
    # 11=comms,12=recreation,13=education,14=restaurants,15=misc
    e3_mapping = [
        ("food_consumption", "CP01", 4),
        ("clothing_consumption", "CP03", 6),
        ("furnishings_consumption", "CP05", 8),
        ("health_consumption", "CP06", 9),
        ("transport_consumption", "CP07", 10),
        ("communication_consumption", "CP08", 11),
        ("recreation_consumption", "CP09", 12),
        ("education_consumption", "CP10", 13),
        ("restaurants_consumption", "CP11", 14),
        ("miscellaneous_consumption", "CP12", 15),
    ]

    out: dict[int, dict[str, float]] = {}
    for yr in TARGET_YEARS:
        row: dict[str, float] = {}
        total_cvm = (
            e3[yr][1] if yr in e3 and isinstance(e3[yr][1], (int, float)) else None
        )
        total_cp = c1.get(yr)

        # The CVM-share fallback only holds near the chained-volume reference year,
        # so it is used solely for 2020+ (post-Eurostat). For 1995-2019 we use the
        # direct Eurostat nominal reading; years Eurostat lacks (pre-1995) get no
        # consumption target rather than a bad CVM-derived one.
        for field, euro_code, e3_col in e3_mapping:
            if yr <= 2019:
                v = estat_get(euro_code, yr)
                if v is not None:
                    row[field] = v / 1000.0  # £m -> £bn
            elif total_cvm and total_cp and yr in e3:
                cvm_val = e3[yr][e3_col]
                if isinstance(cvm_val, (int, float)):
                    row[field] = (
                        (float(cvm_val) / total_cvm) * total_cp / 1000.0
                    )  # £m -> £bn

        # Alcohol and tobacco are one COICOP level-1 division (CP02); target the
        # combined line, not the sub-split. Eurostat nominal for 1995-2019,
        # CVM-share of CP02 × current-prices total for 2020+.
        if yr <= 2019:
            cp02 = estat_get("CP02", yr)
            if cp02:
                row["alcohol_and_tobacco_consumption"] = cp02 / 1000.0
        elif total_cvm and total_cp and yr in e3:
            cp02_cvm = e3[yr][5]
            if isinstance(cp02_cvm, (int, float)):
                row["alcohol_and_tobacco_consumption"] = (
                    (float(cp02_cvm) / total_cvm) * total_cp / 1000.0
                )  # £bn

        # Housing (CP04) excludes owner-occupier imputed rentals (CP042), which are
        # a national-accounts construct, not spending households actually report.
        # Target CP04 - CP042. For 2020+ the E3 CVM sheet has no imputed-rentals
        # sub-line, so scale the CVM housing total by the latest Eurostat actual
        # (non-imputed) share of CP04.
        if yr <= 2019:
            cp04 = estat_get("CP04", yr)
            cp042 = estat_get("CP042", yr)
            if cp04 is not None and cp042 is not None:
                row["housing_water_electricity_consumption"] = (cp04 - cp042) / 1000.0
        elif total_cvm and total_cp and yr in e3:
            cp04_cvm = e3[yr][7]
            cp04_19 = estat_get("CP04", 2019)
            cp042_19 = estat_get("CP042", 2019)
            if isinstance(cp04_cvm, (int, float)) and cp04_19 and cp042_19 is not None:
                actual_share = (cp04_19 - cp042_19) / cp04_19
                row["housing_water_electricity_consumption"] = (
                    (float(cp04_cvm) / total_cvm) * total_cp * actual_share / 1000.0
                )  # £bn

        out[yr] = row

    return out


# ── DWP caseloads ──────────────────────────────────────────────────────────

# Each entry: (variable, sheet, header_row, data_row).
# header_row carries the fiscal-year column labels ("YYYY/YY Outturn");
# data_row holds the caseload total (in thousands) for that benefit.
_CASELOAD_SPECS = [
    ("attendance_allowance", "Disability benefits", 50, 74),
    ("carers_allowance", "Carers Allowance", 14, 15),
    ("disability_living_allowance", "Disability benefits", 50, 55),
    ("esa_income_related", "Incapacity benefits", 119, 129),
    ("housing_benefit", "Housing benefits", 122, 123),
    ("income_support", "Income Support", 90, 91),
    ("jsa_income_based", "Unemployment benefits", 24, 33),
    ("personal_independence_payment", "Disability benefits", 50, 65),
    ("pension_credit", "Pension Credit", 18, 19),
    ("state_pension", "State Pension", 28, 29),
    ("universal_credit", "Universal Credit and equivalent", 46, 53),
]


def _caseload_year_cols(ws, header_row: int) -> dict[int, int]:
    """Map fiscal-year start -> column index from a caseload header row."""
    out: dict[int, int] = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        m = re.search(r"(\d{4})/\d{2}", str(v)) if v else None
        if m:
            out[int(m.group(1))] = c
    return out


def load_dwp_caseloads(path: Path) -> dict[int, dict[str, float]]:
    """Return {fiscal_year_start: {variable: claimant_count}} from the DWP xlsx.

    Caseload totals are stored in thousands; multiply by 1000 for raw counts.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)

    out: dict[int, dict[str, float]] = {}
    for variable, sheet, header_row, data_row in _CASELOAD_SPECS:
        ws = wb[sheet]
        yr_cols = _caseload_year_cols(ws, header_row)
        for yr, col in yr_cols.items():
            v = ws.cell(row=data_row, column=col).value
            if isinstance(v, (int, float)) and v != 0:
                out.setdefault(yr, {})[variable] = float(v) * 1000.0

    return out


# ── Carer's Allowance cases-in-payment ───────────────────────────────────────

# The OBR dwp.xlsx "Carers Allowance" sheet reports CA *entitled* cases, which
# include ~400k people entitled but not paid (the State-Pension overlap rule pays
# the higher benefit instead). CA *expenditure* is paid-only, so pairing it with
# the entitled headcount implies a per-recipient mean ~18% below what the engine
# assigns. Stat-Xplore's CA_In_Payment_New gives the paid-only caseload, which is
# the definition consistent with the expenditure; we use it to override the CA
# caseload target. The DB only starts May 2018, so earlier fiscal years fall back
# to the OBR entitled count scaled by the earliest observed paid share.
_SX_CA_DB = "str:database:CA_In_Payment_New"
_SX_CA_COUNT = "str:count:CA_In_Payment_New:V_F_CA_In_Payment_New"
_SX_CA_DATE = "str:field:CA_In_Payment_New:F_CA_QTR_New:DATE_NAME"
_SX_CA_CACHE = RAW_DIR / "dwp" / "ca_in_payment.json"


def load_ca_in_payment() -> dict[int, float]:
    """Return {fiscal_year_start: mean cases-in-payment} for Carer's Allowance.

    Pulls every quarterly snapshot CA_In_Payment_New publishes (one query, all
    dates), maps each YYYYMM to its UK fiscal year (Apr–Mar: months 1–3 belong to
    the prior fiscal year), and averages the quarters within each year. Cached to
    `_SX_CA_CACHE`; a present cache is reused without hitting the API.
    """
    if _SX_CA_CACHE.exists():
        cached = json.loads(_SX_CA_CACHE.read_text())
        return {int(y): float(v) for y, v in cached.items()}

    key = os.environ.get("STAT_XPLORE_API_KEY")
    if not key:
        raise SystemExit(
            "STAT_XPLORE_API_KEY not set and no cache at "
            f"{_SX_CA_CACHE}; cannot fetch CA cases in payment."
        )
    headers = {"APIKey": key, "Content-Type": "application/json"}

    query = {
        "database": _SX_CA_DB,
        "measures": [_SX_CA_COUNT],
        "dimensions": [[_SX_CA_DATE]],
    }
    resp = requests.post(f"{_SX_BASE}/table", headers=headers, json=query, timeout=120)
    resp.raise_for_status()
    data = resp.json()

    yyyymm = [int(i["uris"][0].split(":")[-1]) for i in data["fields"][0]["items"]]
    values = data["cubes"][_SX_CA_COUNT]["values"]
    by_fy: dict[int, list[float]] = {}
    for ym, v in zip(yyyymm, values):
        if v is None:
            continue
        cal_year, month = ym // 100, ym % 100
        fy = cal_year - 1 if month <= 3 else cal_year
        by_fy.setdefault(fy, []).append(float(v))
    out = {fy: sum(vs) / len(vs) for fy, vs in by_fy.items()}
    _SX_CA_CACHE.parent.mkdir(parents=True, exist_ok=True)
    _SX_CA_CACHE.write_text(json.dumps(out))
    return out


# ── FRS-grossed population targets ────────────────────────────────────────────

_AGE_BANDS = [
    (0, 16),
    (16, 25),
    (25, 35),
    (35, 45),
    (45, 55),
    (55, 65),
    (65, 75),
    (75, None),
]


def build_population_targets(years: list[int]) -> list[dict]:
    """Population control totals grossed up from the FRS by its own household
    weights: total counts of households/people/benunits, plus household count by
    region, person count by sex, and person count by age band. These pin the
    survey universe so reweighting can't drift the totals away from the FRS.
    """
    frs_base = REPO_ROOT / "data" / "clean" / "frs"
    label = "FRS grossed population"
    out: list[dict] = []
    for yr in years:
        ydir = frs_base / str(yr)
        if not (ydir / "households.csv").exists():
            # Silently missing population controls let calibration drift the
            # grossed totals by whole percents — make the gap loud.
            print(
                f"WARNING: no clean FRS at {ydir} — population targets for "
                f"{yr} skipped; calibration for {yr} will be unpinned"
            )
            continue
        hh = pd.read_csv(ydir / "households.csv")
        persons = pd.read_csv(ydir / "persons.csv")
        benunits = pd.read_csv(ydir / "benunits.csv")
        w = hh.set_index("household_id")["weight"].astype(float)

        person_w = persons["household_id"].map(w).to_numpy()
        benunit_w = benunits["household_id"].map(w).to_numpy()

        def add(name, entity, value, flt=None):
            out.append(
                {
                    "name": f"{name}_{yr}",
                    "variable": "household_id",
                    "entity": entity,
                    "aggregation": "count",
                    "filter": flt,
                    "benunit_filter": None,
                    "value": round(float(value), 0),
                    "source": label,
                    "year": yr,
                    "holdout": False,
                }
            )

        add("population_households", "household", w.sum())
        add("population_people", "person", person_w.sum())
        add("population_benunits", "benunit", benunit_w.sum())

        for region in sorted(hh["region"].dropna().unique()):
            val = w[hh.set_index("household_id")["region"] == region].sum()
            slug = re.sub(r"[^a-z0-9]+", "_", str(region).lower()).strip("_")
            add(
                f"population_households_region_{slug}",
                "household",
                val,
                {"variable": "region", "eq": region},
            )

        for sex in sorted(persons["gender"].dropna().unique()):
            val = person_w[(persons["gender"] == sex).to_numpy()].sum()
            add(
                f"population_people_sex_{sex}",
                "person",
                val,
                {"variable": "gender", "eq": sex},
            )

        age = persons["age"].to_numpy()
        for lo, hi in _AGE_BANDS:
            mask = age >= lo
            if hi is not None:
                mask &= age < hi
            val = person_w[mask].sum()
            tag = f"{lo}_{hi}" if hi is not None else f"{lo}_plus"
            add(
                f"population_people_age_{tag}",
                "person",
                val,
                {"variable": "age", "min": lo, "max": hi},
            )
    return out


# ── UC award-amount distribution targets ──────────────────────────────────────

# Stat-Xplore "Households on Universal Credit" (UC_Households), V_F_UC_HOUSEHOLDS
# count, broken down by Monthly Award Amount band, November snapshot each year.
# Pulled live from the open-data REST API at build time (key in the env var
# STAT_XPLORE_API_KEY); the response is cached to data/raw so offline rebuilds
# reuse it. The band axis has 28 values: index 0 = no payment; 1..15 = £0.01-100
# ... £1400.01-1500 per month; 16 = £1500.01+ catch-all (used to Nov 2021);
# 17..27 = £1500.01-1600 ... £2500.01+ fine split (from Nov 2022, with index 16
# then zero). We collapse everything ≥£1500/mo into one band, so the top band =
# index 16 + sum(17..27), robust to that mid-series split.
_SX_BASE = "https://stat-xplore.dwp.gov.uk/webapi/rest/v1"
_SX_DB = "str:database:UC_Households"
_SX_COUNT = "str:count:UC_Households:V_F_UC_HOUSEHOLDS"
_SX_DATE = "str:field:UC_Households:F_UC_HH_DATE:DATE_NAME"
_SX_DATE_VS = "str:valueset:UC_Households:F_UC_HH_DATE:DATE_NAME:C_UC_HH_DATE"
_SX_BAND = "str:field:UC_Households:V_F_UC_HOUSEHOLDS:hnpayment_band"
_SX_CACHE = RAW_DIR / "dwp" / "uc_households_award_bands.json"


def load_uc_award_band_counts() -> dict[int, list[int]]:
    """Return {calendar_year: [28 band counts]} for every November snapshot the
    Stat-Xplore UC_Households table publishes.

    Discovers the available November months from the date valueset, then pulls
    the count by award-amount band for those months in one query. The result is
    cached to `_SX_CACHE`; a present cache is reused without hitting the API.
    """
    if _SX_CACHE.exists():
        cached = json.loads(_SX_CACHE.read_text())
        return {int(y): v for y, v in cached.items()}

    key = os.environ.get("STAT_XPLORE_API_KEY")
    if not key:
        raise SystemExit(
            "STAT_XPLORE_API_KEY not set and no cache at "
            f"{_SX_CACHE}; cannot fetch UC_Households award bands."
        )
    headers = {"APIKey": key, "Content-Type": "application/json"}

    schema = requests.get(
        f"{_SX_BASE}/schema/{_SX_DATE_VS}", headers=headers, timeout=60
    )
    schema.raise_for_status()
    nov_ids = [c["id"] for c in schema.json()["children"] if c["id"].endswith("11")]

    query = {
        "database": _SX_DB,
        "measures": [_SX_COUNT],
        "dimensions": [[_SX_DATE], [_SX_BAND]],
        "recodes": {_SX_DATE: {"map": [[m] for m in nov_ids], "total": False}},
    }
    resp = requests.post(f"{_SX_BASE}/table", headers=headers, json=query, timeout=120)
    resp.raise_for_status()
    data = resp.json()

    months = [int(i["labels"][0].split()[-1]) for i in data["fields"][0]["items"]]
    values = data["cubes"][_SX_COUNT]["values"]
    out = {
        yr: [int(v) if v is not None else 0 for v in row]
        for yr, row in zip(months, values)
    }
    _SX_CACHE.parent.mkdir(parents=True, exist_ok=True)
    _SX_CACHE.write_text(json.dumps(out))
    return out


def _extend_by_caseload(counts_by_year, years, uc_caseload, scale):
    """Carry a per-year UC count series past its last observed snapshot.

    Stat-Xplore UC_Households stops at the Nov 2023 snapshot, so for requested
    years beyond it we hold the within-caseload composition fixed and scale the
    last observed year by the UC caseload's relative growth (the same DWP
    trajectory that drives the headcount). `scale(value, ratio)` applies the
    growth factor to whatever value shape the caller stores (a count, or a list
    of band counts).
    """
    out = dict(counts_by_year)
    if not counts_by_year:
        return out
    anchor = max(counts_by_year)
    base = uc_caseload.get(anchor)
    if not base:
        return out
    for yr in sorted(set(years)):
        if yr <= anchor or yr in out or yr not in uc_caseload:
            continue
        out[yr] = scale(counts_by_year[anchor], uc_caseload[yr] / base)
    return out


def build_uc_award_band_targets(years: list[int]) -> list[dict]:
    """Count of UC benefit units by monthly award-amount band, calibrated to the
    DWP Households-on-UC distribution. The simulation's annual UC entitlement is
    binned into monthly £100 bands (annual edge = monthly × 12), pinning the
    *shape* of the award distribution rather than just the total caseload. The
    nil-award band is dropped (a £0 filter on UC alone can't distinguish non-
    recipients) and everything ≥£1500/mo is collapsed into one top band.

    Only observed Nov snapshots are targeted: a band shape carried past the
    last snapshot is blind to policy changes that move awards between bands
    (e.g. the April 2026 two-child-limit repeal), so calibration would reweight
    against the reform. Beyond the snapshot horizon the UC £ total is trained
    instead (its holdout is lifted there — see HOLDOUT_NAMES).
    """
    band_counts = load_uc_award_band_counts()
    label = "DWP Stat-Xplore UC_Households award bands (Nov snapshot)"
    out: list[dict] = []
    for yr in years:
        counts = band_counts.get(yr)
        if counts is None:
            continue

        def add(name, value, lo_monthly, hi_monthly):
            flt = {
                "variable": "universal_credit",
                "min": max(lo_monthly * 12.0, 0.01),
                "max": None if hi_monthly is None else hi_monthly * 12.0,
            }
            out.append(
                {
                    "name": f"{name}_{yr}",
                    "variable": "universal_credit",
                    "entity": "benunit",
                    "aggregation": "count",
                    "filter": flt,
                    "benunit_filter": None,
                    "value": round(float(value), 0),
                    "source": label,
                    "year": yr,
                    "holdout": False,
                }
            )

        for b in range(1, 16):
            lo, hi = (b - 1) * 100, b * 100
            add(f"uc_award_band_{lo}_{hi}", counts[b], lo, hi)
        add("uc_award_band_1500_plus", counts[16] + sum(counts[17:28]), 1500, None)
    return out


# ── UC element-entitlement counts ─────────────────────────────────────────────

# Stat-Xplore UC_Households yes/no entitlement fields, November snapshot. Each
# maps to a benunit-level UC element the engine emits (uc_<key>_element); the
# target is the count of UC benunits with that element > 0. LCW is modelled only
# as LCWRA in the engine, so the limited-capability target uses the LCWRA value,
# not the LCW-or-LCWRA total.
_SX_ELEMENT_FIELDS = {  # engine element variable -> (Stat-Xplore field id, "yes" value index)
    "uc_carer_element": ("HCCARER_ENTITLEMENT", "1"),
    "uc_child_element": ("HCCHILD_ENTITLEMENT", "1"),
    "uc_disabled_child_element": ("HCDISABLED_CHILD_ENTITLEMENT", "1"),
    "uc_housing_element": ("TENURE", "1"),
    "uc_lcwra_element": ("HCLCW_ENTITLEMENT", "3"),  # 3 = LCWRA (not 1 = LCW-or-LCWRA)
}
# The disabled-child element is fetched (for the diagnostic record) but not turned
# into a calibration target. The FRS frame carries only ~100 UC benefit units with
# a disabled child (1.7k grossed in 2019, 9k in 2023), far below the DWP counts
# (76k / 318k), so reaching the target needs the 10x weight clamp and still falls
# 11-36% short — a thin-cell survey limit, not an optimiser failure. Chasing it
# would distort the housing/child/LCWRA elements those same records feed, so we
# emit the other four elements and leave disabled-child out of the loss.
_UC_ELEMENT_TARGETS = [
    v for v in _SX_ELEMENT_FIELDS if v != "uc_disabled_child_element"
]
_SX_ELEMENT_CACHE = RAW_DIR / "dwp" / "uc_households_elements.json"


def load_uc_element_counts() -> dict[str, dict[int, int]]:
    """Return {engine_element_var: {calendar_year: count}} of UC households with
    each entitlement, for every November snapshot UC_Households publishes.

    One query per element field (the "yes"/LCWRA value picked out), all Novembers
    at once. Cached to `_SX_ELEMENT_CACHE`; a present cache is reused offline.
    """
    if _SX_ELEMENT_CACHE.exists():
        cached = json.loads(_SX_ELEMENT_CACHE.read_text())
        return {var: {int(y): c for y, c in yrs.items()} for var, yrs in cached.items()}

    key = os.environ.get("STAT_XPLORE_API_KEY")
    if not key:
        raise SystemExit(
            "STAT_XPLORE_API_KEY not set and no cache at "
            f"{_SX_ELEMENT_CACHE}; cannot fetch UC_Households elements."
        )
    headers = {"APIKey": key, "Content-Type": "application/json"}

    schema = requests.get(
        f"{_SX_BASE}/schema/{_SX_DATE_VS}", headers=headers, timeout=60
    )
    schema.raise_for_status()
    nov_ids = [c["id"] for c in schema.json()["children"] if c["id"].endswith("11")]

    out: dict[str, dict[int, int]] = {}
    for var, (field, want) in _SX_ELEMENT_FIELDS.items():
        field_id = f"str:field:UC_Households:V_F_UC_HOUSEHOLDS:{field}"
        query = {
            "database": _SX_DB,
            "measures": [_SX_COUNT],
            "dimensions": [[_SX_DATE], [field_id]],
            "recodes": {_SX_DATE: {"map": [[m] for m in nov_ids], "total": False}},
        }
        resp = requests.post(
            f"{_SX_BASE}/table", headers=headers, json=query, timeout=120
        )
        resp.raise_for_status()
        data = resp.json()
        months = [int(i["labels"][0].split()[-1]) for i in data["fields"][0]["items"]]
        # Locate the wanted value's column on the entitlement axis by its value id.
        col_ids = [i["uris"][0].split(":")[-1] for i in data["fields"][1]["items"]]
        ci = col_ids.index(want)
        values = data["cubes"][_SX_COUNT]["values"]
        out[var] = {
            yr: int(row[ci]) if row[ci] is not None else 0
            for yr, row in zip(months, values)
        }
    _SX_ELEMENT_CACHE.parent.mkdir(parents=True, exist_ok=True)
    _SX_ELEMENT_CACHE.write_text(json.dumps(out))
    return out


def build_uc_element_targets(
    years: list[int], uc_caseload: dict[int, float]
) -> list[dict]:
    """Count of UC benefit units claiming each UC element, calibrated to the DWP
    Households-on-UC entitlement breakdown. Each element maps to the benunit-level
    uc_<key>_element the engine emits; the target counts benunits with it > 0.
    Pins the *composition* of the UC caseload (how many get the child, housing,
    carer, disability and LCWRA additions) on top of the headcount and award-band
    shape. Years past the last Nov snapshot (2023) hold each element's share of
    the caseload fixed and grow with the UC caseload trajectory.
    """
    raw_counts = load_uc_element_counts()
    element_counts = {
        var: _extend_by_caseload(
            raw_counts[var], years, uc_caseload, lambda v, r: v * r
        )
        for var in _UC_ELEMENT_TARGETS
    }
    label = "DWP Stat-Xplore UC_Households elements (Nov snapshot)"
    out: list[dict] = []
    for yr in years:
        for var in _UC_ELEMENT_TARGETS:
            value = element_counts[var].get(yr)
            if value is None:
                continue
            out.append(
                {
                    "name": f"{var}_claimants_{yr}",
                    "variable": var,
                    "entity": "benunit",
                    "aggregation": "count_nonzero",
                    "filter": None,
                    "benunit_filter": None,
                    "value": round(float(value), 0),
                    "source": label,
                    "year": yr,
                    "holdout": False,
                }
            )
    return out


# ── UC in-work caseload ───────────────────────────────────────────────────────

# Stat-Xplore "People on Universal Credit" (UC_Monthly) employment indicator,
# banded to in-work vs not. This database is frozen at August 2021 (latest
# November snapshot Nov 2019 once the COVID surge is excluded), so we take the
# Nov-2019 in-work people count as the anchor and carry it forward by the UC
# caseload's relative growth (the same DWP-trajectory series used for the
# headcount). Mapping: in-work people on UC = persons in a UC benefit unit with
# positive earned income (children carry zero earnings and drop out), so the
# engine target is a person-level count_nonzero of earned income filtered to UC
# recipients.
_SX_UCM_DB = "str:database:UC_Monthly"
_SX_UCM_COUNT = "str:count:UC_Monthly:V_F_UC_CASELOAD_FULL"
_SX_UCM_DATE = "str:field:UC_Monthly:F_UC_DATE:DATE_NAME"
_SX_UCM_DATE_VS = "str:valueset:UC_Monthly:F_UC_DATE:DATE_NAME:C_UC_DATE"
_SX_UCM_EMP_VS = "str:valueset:UC_Monthly:V_F_UC_CASELOAD_FULL:EMPLOYMENT_CODE:C_UC_EMPLOYMENT_2023_BAND"
_SX_UCM_CACHE = RAW_DIR / "dwp" / "uc_monthly_inwork.json"
# Anchor on Nov 2019: the latest pre-pandemic November (Aug-2021 freeze means
# Nov 2020 is the last snapshot, but its 49% in-work share is COVID-inflated).
UC_INWORK_ANCHOR = 2019


def load_uc_inwork_counts() -> dict[int, int]:
    """Return {calendar_year: in-work people on UC} for each November snapshot
    UC_Monthly publishes (in-work = the "In employment (PAYE) or self-employment"
    band). Cached to `_SX_UCM_CACHE`; a present cache is reused offline.
    """
    if _SX_UCM_CACHE.exists():
        cached = json.loads(_SX_UCM_CACHE.read_text())
        return {int(y): v for y, v in cached.items()}

    key = os.environ.get("STAT_XPLORE_API_KEY")
    if not key:
        raise SystemExit(
            "STAT_XPLORE_API_KEY not set and no cache at "
            f"{_SX_UCM_CACHE}; cannot fetch UC_Monthly in-work counts."
        )
    headers = {"APIKey": key, "Content-Type": "application/json"}

    schema = requests.get(
        f"{_SX_BASE}/schema/{_SX_UCM_DATE_VS}", headers=headers, timeout=60
    )
    schema.raise_for_status()
    nov_ids = [c["id"] for c in schema.json()["children"] if c["id"].endswith("11")]

    query = {
        "database": _SX_UCM_DB,
        "measures": [_SX_UCM_COUNT],
        "dimensions": [[_SX_UCM_DATE], [_SX_UCM_EMP_VS]],
        "recodes": {_SX_UCM_DATE: {"map": [[m] for m in nov_ids], "total": False}},
    }
    resp = requests.post(f"{_SX_BASE}/table", headers=headers, json=query, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    months = [int(i["labels"][0].split()[-1]) for i in data["fields"][0]["items"]]
    emp_labels = [i["labels"][0] for i in data["fields"][1]["items"]]
    in_work_col = next(
        i for i, lbl in enumerate(emp_labels) if lbl.startswith("In employment")
    )
    values = data["cubes"][_SX_UCM_COUNT]["values"]
    out = {
        yr: int(row[in_work_col])
        for yr, row in zip(months, values)
        if row[in_work_col] is not None and row[in_work_col] > 0
    }
    _SX_UCM_CACHE.parent.mkdir(parents=True, exist_ok=True)
    _SX_UCM_CACHE.write_text(json.dumps(out))
    return out


def build_uc_inwork_targets(
    years: list[int], uc_caseload: dict[int, float]
) -> list[dict]:
    """Count of people on UC who are in work, calibrated to the DWP UC_Monthly
    employment indicator. Actuals run to the anchor (Nov 2019); later years take
    the anchor scaled by the UC caseload's relative growth, so the in-work count
    tracks the headcount trajectory while holding the in-work share at its last
    pre-pandemic level. Maps to a person-level count of earned income > 0 filtered
    to UC recipients.
    """
    inwork = load_uc_inwork_counts()
    label = "DWP Stat-Xplore UC_Monthly in-work caseload"
    actuals = {y: v for y, v in inwork.items() if y <= UC_INWORK_ANCHOR}
    if not actuals:
        return []
    anchor_year = max(actuals)
    anchor = actuals[anchor_year]
    series: dict[int, float] = dict(actuals)
    base_caseload = uc_caseload.get(anchor_year)
    if base_caseload:
        for yr in sorted(y for y in uc_caseload if y > anchor_year):
            series[yr] = anchor * uc_caseload[yr] / base_caseload

    out: list[dict] = []
    for yr in years:
        value = series.get(yr)
        if value is None:
            continue
        out.append(
            {
                "name": f"uc_in_work_people_{yr}",
                "variable": "earned_income",
                "entity": "person",
                "aggregation": "count_nonzero",
                "filter": {"variable": "universal_credit", "min": 0.01, "max": None},
                "benunit_filter": None,
                "value": round(float(value), 0),
                "source": label,
                "year": yr,
                "holdout": False,
            }
        )
    return out


# ── Forecast-year targets (uprated from the latest real year) ─────────────────

# Per-source uprating index for forecast years. Each target grows at the rate of
# the microdata variable that generates it, so forecast-year calibration stays a
# nudge rather than fighting the uprating. Counts (population/caseloads/labour)
# track the population index; expenditure and consumption track CPI; tax receipts
# track the base of the tax (earnings for income tax/NI/SDLT, CPI for VAT, GDP per
# capita for CGT). SPI distribution targets are handled separately.
_FORECAST_INDEX = {
    "FRS grossed population": "population",
    "OBR EFO economy table 1.6 (labour market)": "population",
    "DWP Stat-Xplore UC_Households award bands (Nov snapshot)": "population",
    "DWP Stat-Xplore UC_Households elements (Nov snapshot)": "population",
    "Eurostat/ONS HHFCE COICOP": "cpi",
}
# DWP benefit expenditure carries both £ sums (CPI) and claimant counts
# (population), distinguished by aggregation. HMRC receipts vary by tax.
_HMRC_RECEIPT_INDEX = {
    "income_tax": "earnings",
    "employee_ni": "earnings",
    "employer_ni": "earnings",
    "stamp_duty": "earnings",
    "vat": "cpi",
    "capital_gains_tax": "gdp_pc",
}


def _forecast_index_for(t: dict) -> str:
    src = t["source"]
    if src == "HMRC receipts":
        return _HMRC_RECEIPT_INDEX[t["variable"]]
    if src == "DWP benefit expenditure":
        return "population" if t["aggregation"] == "count_nonzero" else "cpi"
    return _FORECAST_INDEX[src]


def _retag_name(name: str, base_year: int, forecast_year: int) -> str:
    suffix = f"_{base_year}"
    return name[: -len(suffix)] + f"_{forecast_year}" if name.endswith(suffix) else name


def build_forecast_targets(
    real_targets: list[dict], forecast_years: list[int]
) -> list[dict]:
    """Project the latest real year's targets onto OBR forecast years.

    Non-SPI targets are scaled by the cumulative growth of their source index
    (counts by population, expenditure/consumption by CPI, receipts by their tax
    base). SPI distribution targets uprate their income thresholds and amounts by
    earnings growth while holding counts fixed (the band shape is assumed stable).
    """
    base = LATEST_REAL_YEAR
    spi_labels = (
        "HMRC SPI Table 3.6 (income distribution)",
        "HMRC SPI Table 3.7 (investment income distribution)",
    )
    # The UC Stat-Xplore breakdowns (award bands, elements, in-work) are built
    # directly for every year from the caseload-anchored series, so they must not
    # be re-projected here (would duplicate the target).
    _uc_direct = (
        "DWP Stat-Xplore UC_Households award bands (Nov snapshot)",
        "DWP Stat-Xplore UC_Households elements (Nov snapshot)",
        "DWP Stat-Xplore UC_Monthly in-work caseload",
    )
    base_targets = [
        t for t in real_targets if t["year"] == base and t["source"] not in _uc_direct
    ]
    out: list[dict] = []

    for yr in forecast_years:
        for t in base_targets:
            if t["source"] in spi_labels:
                ef = cumulative_factor(base, yr, "earnings")
                flt = t["filter"]
                new_flt = {
                    "variable": flt["variable"],
                    "min": None if flt["min"] is None else round(flt["min"] * ef, 0),
                    "max": None if flt["max"] is None else round(flt["max"] * ef, 0),
                }
                # Counts held fixed; amounts grow with earnings. Re-tag the band
                # in the name to the uprated lower threshold.
                tag = 0 if new_flt["min"] is None else int(new_flt["min"])
                stem = t["name"].rsplit("_", 2)[0]  # spi_<key>_<count|amount>
                value = (
                    t["value"]
                    if t["aggregation"] == "count_nonzero"
                    else round(t["value"] * ef, 0)
                )
                nt = dict(t)
                nt.update(
                    name=f"{stem}_{tag}_{yr}", filter=new_flt, value=value, year=yr
                )
                out.append(nt)
            else:
                f = cumulative_factor(base, yr, _forecast_index_for(t))
                nt = dict(t)
                nt.update(
                    name=_retag_name(t["name"], base, yr),
                    value=round(t["value"] * f, 0),
                    year=yr,
                )
                out.append(nt)
    return out


# ── Assemble targets ─────────────────────────────────────────────────────────


def build_targets(years: list[int]) -> list[dict]:
    hmrc_path = RAW_DIR / "hmrc" / "NS_Table.ods"
    dwp_path = RAW_DIR / "obr" / "dwp.xlsx"
    qna_path = RAW_DIR / "ons" / "qna_hhfce.xlsx"

    _download(HMRC_ODS_URL, hmrc_path)
    _download(ONS_QNA_URL, qna_path)

    hmrc = load_hmrc_receipts(hmrc_path)
    # Backfill pre-2006 tax (HMRC bulletin starts 2005-06) from the OBR databank,
    # which covers 1999-00 onwards. HMRC takes priority where both exist.
    databank = load_obr_databank_receipts(RAW_DIR / "obr" / "pf_databank.xlsx")
    for fy, vals in databank.items():
        hmrc.setdefault(fy, vals)
    dwp = load_dwp_benefits(dwp_path)
    coicop = load_coicop(qna_path)
    caseloads = load_dwp_caseloads(dwp_path)
    # Override the CA caseload (OBR entitled) with the Stat-Xplore cases-in-payment
    # series, so the headcount matches the paid-only expenditure target. Years
    # before the in-payment DB starts (2018) keep the OBR entitled count scaled by
    # the earliest observed paid share; forecast years past the latest snapshot
    # are left to the generic forecast uprating of the corrected level.
    ca_paid = load_ca_in_payment()
    if ca_paid:
        first_paid_fy = min(ca_paid)
        ca_entitled = {
            yr: d["carers_allowance"]
            for yr, d in caseloads.items()
            if "carers_allowance" in d
        }
        paid_share = (
            ca_paid[first_paid_fy] / ca_entitled[first_paid_fy]
            if ca_entitled.get(first_paid_fy)
            else 1.0
        )
        for yr in ca_entitled:
            if yr in ca_paid:
                caseloads[yr]["carers_allowance"] = ca_paid[yr]
            elif yr < first_paid_fy:
                caseloads[yr]["carers_allowance"] = ca_entitled[yr] * paid_share
    # UC caseload: use the Stat-Xplore UC_Households paying-household count (the
    # award-band sum over index 1+, dropping the nil-award band 0) as the actual
    # series up to its latest November snapshot, then carry it forward by the
    # *relative* growth of the DWP outturn-and-forecast caseload table. The DWP
    # table carries no UC before 2019 (UC was pilot-scale) and reports it on a
    # different basis, so taking only its year-on-year growth ratios — anchored
    # on the latest Stat-Xplore actual — keeps the level survey-consistent while
    # following the published forecast trajectory (legacy-benefit migration keeps
    # UC growing well above population through the late 2020s).
    uc_band_counts = load_uc_award_band_counts()
    uc_paying = {yr: float(sum(c[1:])) for yr, c in uc_band_counts.items()}
    uc_anchor_year = max(uc_paying)
    dwp_uc = {
        yr: d["universal_credit"]
        for yr, d in caseloads.items()
        if "universal_credit" in d
    }
    for yr, val in uc_paying.items():
        caseloads.setdefault(yr, {})["universal_credit"] = val
    anchor = uc_paying[uc_anchor_year]
    for yr in sorted(y for y in dwp_uc if y > uc_anchor_year):
        if dwp_uc.get(uc_anchor_year):
            caseloads.setdefault(yr, {})["universal_credit"] = (
                anchor * dwp_uc[yr] / dwp_uc[uc_anchor_year]
            )
    labour = load_efo_labour_market(RAW_DIR / "obr" / "efo_economy.xlsx")
    spi = load_spi_distributions(RAW_DIR / "hmrc" / "spi")
    spi37 = load_spi_investment(RAW_DIR / "hmrc" / "spi")
    earnings_index = load_efo_earnings_index(RAW_DIR / "obr" / "efo_economy.xlsx")

    # Variable specs: (name_prefix, entity, variable, aggregation, source_key_in_dict, scale)
    # Scale: all final values should be in £ (the raw survey unit is weekly £ per hh,
    # so calibrate target = aggregate_£_per_year × 52  — handled by the matrix builder
    # summing over households. But since benefit/tax values on Person/BenUnit are annual £,
    # the target should be the national annual total in £.)
    INCOME_TAX_SPECS = [
        ("income_tax_total", "person", "income_tax", "sum", "hmrc", "income_tax"),
        ("employee_ni_total", "person", "employee_ni", "sum", "hmrc", "employee_ni"),
        # Employer NI is intentionally not a calibration target: the engine
        # computes gross statutory liability (15% above the secondary threshold)
        # with no reliefs, while the HMRC receipts figure is net of the
        # Employment Allowance, under-21/apprentice exemptions and the per-job
        # (not per-person) threshold. The ~28% gap is structural, so reweighting
        # cannot close it and only distorts the weight vector chasing it.
        # Capital gains tax is intentionally not a calibration target: the FRS
        # carries no realised-gains data, so the engine predicts ~£0 and the
        # target has no survey representation to reweight against — it only ever
        # showed as an untrained -100% miss.
        ("vat_total", "household", "vat", "sum", "hmrc", "vat"),
        # Stamp duty is intentionally not a calibration target: the engine
        # under-predicts receipts by ~15-19% every year (missing higher-rate
        # surcharges on additional/non-resident dwellings, plus survey under-
        # capture of high-value transactions). The gap is structural and stable,
        # so reweighting only distorts the weight vector chasing it.
    ]
    BENEFIT_SPECS = [
        (
            "child_benefit_total",
            "benunit",
            "child_benefit",
            "sum",
            "dwp",
            "child_benefit",
        ),
        (
            "attendance_allowance_total",
            "person",
            "attendance_allowance",
            "sum",
            "dwp",
            "attendance_allowance",
        ),
        (
            "carers_allowance_total",
            "person",
            "carers_allowance",
            "sum",
            "dwp",
            "carers_allowance",
        ),
        (
            "disability_living_allowance_total",
            "person",
            "disability_living_allowance",
            "sum",
            "dwp",
            "disability_living_allowance",
        ),
        (
            "personal_independence_payment_total",
            "person",
            "personal_independence_payment",
            "sum",
            "dwp",
            "personal_independence_payment",
        ),
        (
            "esa_income_related_total",
            "benunit",
            "esa_income_related",
            "sum",
            "dwp",
            "esa_income_related",
        ),
        (
            "housing_benefit_total",
            "benunit",
            "housing_benefit",
            "sum",
            "dwp",
            "housing_benefit",
        ),
        (
            "income_support_total",
            "benunit",
            "income_support",
            "sum",
            "dwp",
            "income_support",
        ),
        # Income-based JSA expenditure is intentionally not a calibration target,
        # for the same reason as its caseload (below): the benefit is wound down
        # into UC, and the engine under-predicts spend by 64-97% every year. The
        # gap is structural — the FRS frame can't carry the residual legacy
        # caseload — so reweighting only thrashes the weight vector chasing it.
        (
            "pension_credit_total",
            "benunit",
            "pension_credit",
            "sum",
            "dwp",
            "pension_credit",
        ),
        (
            "state_pension_total",
            "person",
            "state_pension",
            "sum",
            "dwp",
            "state_pension",
        ),
        (
            "universal_credit_total",
            "benunit",
            "universal_credit",
            "sum",
            "dwp",
            "universal_credit",
        ),
        ("tax_credits_total", "benunit", "tax_credits", "sum", "dwp", "tax_credits"),
    ]
    COICOP_SPECS = [
        (
            "food_consumption_total",
            "household",
            "food_consumption",
            "sum",
            "coicop",
            "food_consumption",
        ),
        (
            "alcohol_and_tobacco_consumption_total",
            "household",
            "alcohol_and_tobacco_consumption",
            "sum",
            "coicop",
            "alcohol_and_tobacco_consumption",
        ),
        (
            "clothing_consumption_total",
            "household",
            "clothing_consumption",
            "sum",
            "coicop",
            "clothing_consumption",
        ),
        (
            "housing_water_electricity_consumption_total",
            "household",
            "housing_water_electricity_consumption",
            "sum",
            "coicop",
            "housing_water_electricity_consumption",
        ),
        (
            "furnishings_consumption_total",
            "household",
            "furnishings_consumption",
            "sum",
            "coicop",
            "furnishings_consumption",
        ),
        (
            "health_consumption_total",
            "household",
            "health_consumption",
            "sum",
            "coicop",
            "health_consumption",
        ),
        (
            "transport_consumption_total",
            "household",
            "transport_consumption",
            "sum",
            "coicop",
            "transport_consumption",
        ),
        (
            "communication_consumption_total",
            "household",
            "communication_consumption",
            "sum",
            "coicop",
            "communication_consumption",
        ),
        (
            "recreation_consumption_total",
            "household",
            "recreation_consumption",
            "sum",
            "coicop",
            "recreation_consumption",
        ),
        (
            "education_consumption_total",
            "household",
            "education_consumption",
            "sum",
            "coicop",
            "education_consumption",
        ),
        (
            "restaurants_consumption_total",
            "household",
            "restaurants_consumption",
            "sum",
            "coicop",
            "restaurants_consumption",
        ),
        (
            "miscellaneous_consumption_total",
            "household",
            "miscellaneous_consumption",
            "sum",
            "coicop",
            "miscellaneous_consumption",
        ),
    ]

    CASELOAD_SPECS = [
        (
            "attendance_allowance_claimants",
            "person",
            "attendance_allowance",
            "count_nonzero",
            "caseloads",
            "attendance_allowance",
        ),
        (
            "carers_allowance_claimants",
            "person",
            "carers_allowance",
            "count_nonzero",
            "caseloads",
            "carers_allowance",
        ),
        (
            "disability_living_allowance_claimants",
            "person",
            "disability_living_allowance",
            "count_nonzero",
            "caseloads",
            "disability_living_allowance",
        ),
        (
            "personal_independence_payment_claimants",
            "person",
            "personal_independence_payment",
            "count_nonzero",
            "caseloads",
            "personal_independence_payment",
        ),
        (
            "esa_income_related_claimants",
            "benunit",
            "esa_income_related",
            "count_nonzero",
            "caseloads",
            "esa_income_related",
        ),
        (
            "housing_benefit_claimants",
            "benunit",
            "housing_benefit",
            "count_nonzero",
            "caseloads",
            "housing_benefit",
        ),
        (
            "income_support_claimants",
            "benunit",
            "income_support",
            "count_nonzero",
            "caseloads",
            "income_support",
        ),
        # Income-based JSA caseload is intentionally not a calibration target:
        # the benefit is wound down into UC, leaving a residual legacy caseload
        # the FRS frame can't represent, so the engine under-predicts it every
        # year. The expenditure target is dropped for the same reason (see
        # BENEFIT_SPECS above). Reweighting only thrashes the weights chasing it.
        (
            "pension_credit_claimants",
            "benunit",
            "pension_credit",
            "count_nonzero",
            "caseloads",
            "pension_credit",
        ),
        # State-pension caseload is intentionally not a calibration target. Receipt
        # is near-universal among the survey's over-66s (only ~0.2-0.3% carry no
        # state pension from 2021), so the recipient set and the over-State-Pension-
        # age population are almost the same households. The DWP caseload sits ~1%
        # above the FRS-grossed count (it includes overseas pensioners the GB survey
        # frame can't carry), so chasing it fights the FRS age-band population
        # anchors: the optimiser can't lift claimants without pushing the 65-75/75+
        # bands over. We keep the age bands and the state_pension expenditure target
        # (which fits cleanly) and drop the headcount.
        (
            "universal_credit_claimants",
            "benunit",
            "universal_credit",
            "count_nonzero",
            "caseloads",
            "universal_credit",
        ),
    ]

    # is_employed/is_unemployed are 0/1 person flags, so the target value is a
    # headcount, not a £ amount. Tag the aggregation as count_nonzero (not sum)
    # so calibrate.py applies the people sub-threshold floor (1e4) rather than
    # the £50m sum floor — otherwise a 1.5m-person unemployment target is dropped
    # as "sub-threshold £50m" and never trained.
    LABOUR_SPECS = [
        (
            "employed_count",
            "person",
            "is_employed",
            "count_nonzero",
            "labour",
            "employed",
        ),
        (
            "unemployed_count",
            "person",
            "is_unemployed",
            "count_nonzero",
            "labour",
            "unemployed",
        ),
    ]

    all_specs = (
        INCOME_TAX_SPECS + BENEFIT_SPECS + COICOP_SPECS + CASELOAD_SPECS + LABOUR_SPECS
    )
    # Held out of the calibration loss but still scored in the diagnostics.
    #
    # Income Support's expenditure (£280m, DWP Table 1a) and caseload (43k) come
    # from two independent DWP series whose implied per-recipient mean (£6.5k) is
    # 27% below what the engine assigns (£8.3k). Reweighting scales record copies,
    # not per-record amounts, so the optimiser can hit neither without the other
    # drifting; at IS's tiny size it isn't worth distorting the weight vector.
    #
    # Income tax, employee NI and VAT receipts are held out because the engine
    # emits statutory *liability* while the NS_Table targets are HMRC *cash
    # receipts*; liability exceeds receipts by the tax gap (the uncollected /
    # evaded / avoided portion). The engine over-predicts each on an income/
    # consumption base the SPI and COICOP targets already pin to <0.1%, so there
    # is no compositional slack left for reweighting to close the gap (income tax
    # +3.3%, NI +5.2%, VAT +7.3% in 2024 — the published UK tax-gap ordering).
    # This is the same liability-vs-receipts mismatch that already drops employer
    # NI, CGT and stamp duty (see INCOME_TAX_SPECS). Kept visible (not deleted)
    # so the gap stays in the report rather than being silently scaled away.
    #
    # UC and state pension *totals* are held out because each contradicts a finer
    # target set the survey fits cleanly, so reweighting can't hit both. UC: the
    # Stat-Xplore award-band counts already pin the UC distribution shape, so the
    # aggregate £ total can't also be matched once the per-band amounts differ from
    # the engine's. This only applies up to the last observed band snapshot — past
    # it there are no band targets, and the £ total is trained instead. State pension: the total competes with the FRS-grossed 65-75 /
    # 75+ age bands — the same pensioner households carry both, so lifting weights
    # to reach the SP total overshoots the elderly population controls. We keep the
    # distribution (award bands) and the population bands, holding out the totals.
    #
    # Carers' allowance: the in-payment caseload fix (Stat-Xplore) narrowed the
    # count/total mean gap but a residual remains the 10x clamp can't close, so the
    # pair splits it (~±1.3%). We hold out the *claimants* count and keep the £
    # expenditure total, since costings read against CA spend, not headcount.
    HOLDOUT_NAMES = {
        "income_support_total",
        "income_support_claimants",
        "income_tax_total",
        "employee_ni_total",
        "vat_total",
        "universal_credit_total",
        "state_pension_total",
        "carers_allowance_claimants",
    }
    source_map = {
        "hmrc": hmrc,
        "dwp": dwp,
        "coicop": coicop,
        "caseloads": caseloads,
        "labour": labour,
    }
    source_label = {
        "hmrc": "HMRC receipts",
        "dwp": "DWP benefit expenditure",
        "coicop": "Eurostat/ONS HHFCE COICOP",
        "caseloads": "DWP benefit expenditure",
        "labour": "OBR EFO economy table 1.6 (labour market)",
    }

    real_years = [y for y in years if y <= LATEST_REAL_YEAR]
    # Award-band targets stop at the last observed Nov snapshot; past it the UC
    # £ total is trained instead (the shape/total conflict that justified the
    # holdout only exists where band targets are present).
    uc_band_anchor = max(load_uc_award_band_counts())
    targets = []
    for yr in real_years:
        for name, entity, variable, aggregation, source_key, data_key in all_specs:
            data = source_map[source_key].get(yr, {})
            raw = data.get(data_key)
            if raw is None:
                continue
            # caseloads and labour levels are raw counts; everything else is £bn -> £
            val = raw if source_key in ("caseloads", "labour") else raw * 1e9
            holdout = name in HOLDOUT_NAMES
            if name == "universal_credit_total" and yr > uc_band_anchor:
                holdout = False
            targets.append(
                {
                    "name": f"{name}_{yr}",
                    "variable": variable,
                    "entity": entity,
                    "aggregation": aggregation,
                    "filter": None,
                    "benunit_filter": None,
                    "value": round(val, 0),
                    "source": source_label[source_key],
                    "year": yr,
                    "holdout": holdout,
                }
            )

    targets += build_spi_targets(spi, earnings_index, real_years)
    targets += build_spi_investment_targets(spi37, earnings_index, real_years)
    targets += build_population_targets(real_years)
    # The UC Stat-Xplore tables stop at the Nov 2023 snapshot, but UC is still
    # absorbing legacy-benefit migration through the late 2020s, so the band,
    # element and in-work breakdowns are carried forward by the DWP UC caseload
    # trajectory (holding within-caseload composition fixed) rather than left to
    # the generic forecast uprating. They are produced for every requested year
    # here and excluded from the generic forecast pass below.
    uc_caseload = {
        yr: d["universal_credit"]
        for yr, d in caseloads.items()
        if "universal_credit" in d
    }
    targets += build_uc_award_band_targets(years)
    targets += build_uc_element_targets(years, uc_caseload)
    targets += build_uc_inwork_targets(years, uc_caseload)

    forecast_years = [y for y in years if y > LATEST_REAL_YEAR]
    if forecast_years:
        # Forecast targets project the latest real year forward; build that base
        # even if the caller didn't request it.
        base = (
            targets
            if LATEST_REAL_YEAR in real_years
            else build_targets([LATEST_REAL_YEAR])
        )
        targets += build_forecast_targets(base, forecast_years)

    # The generic forecast uprating grows the UC caseload by population, but UC is
    # still absorbing legacy-benefit migration through the late 2020s, so the DWP
    # table projects it well above population. Override the UC caseload target in
    # every year with the Stat-Xplore-anchored, DWP-trajectory series built above.
    for t in targets:
        if (
            t["variable"] == "universal_credit"
            and t["aggregation"] == "count_nonzero"
            and t["year"] in caseloads
            and "universal_credit" in caseloads[t["year"]]
        ):
            t["value"] = round(caseloads[t["year"]]["universal_credit"], 0)

    # DWP-table programmes: the generic forecast pass is policy-blind (CPI /
    # population scaling of the 2024 value), while the DWP outturn-and-forecast
    # table costs each programme directly — including managed migration (legacy
    # spend and caseloads → ~0 by 2026) and the two-child-limit repeal inside
    # UC. Override every DWP-sourced £ total and claimant count with the table
    # value wherever it has one; where a legacy programme drops out of the
    # table (or falls below £0.1bn) in a forecast year, stop training the
    # target rather than chase an extrapolation of a benefit that no longer
    # exists. UC's claimant count keeps the Stat-Xplore-anchored series built
    # above (same trajectory, survey-consistent level).
    _dwp_progs = [
        "universal_credit",
        "tax_credits",
        "esa_income_related",
        "housing_benefit",
        "income_support",
        "jsa_income_based",
        "pension_credit",
        "state_pension",
        "child_benefit",
    ]
    for t in targets:
        prog = t["variable"]
        if prog not in _dwp_progs:
            continue
        # Only unfiltered national totals: SPI band targets and the UC award
        # band/element breakdowns share these variable names but carry filters.
        if t.get("filter") is not None or t.get("benunit_filter") is not None:
            continue
        yr = t["year"]
        if t["aggregation"] == "sum":
            if prog in dwp.get(yr, {}):
                val = dwp[yr][prog] * 1e9
                t["value"] = round(max(val, 0.0), 0)
                if val < 0.1e9:
                    t["holdout"] = True
            elif yr > LATEST_REAL_YEAR:
                t["holdout"] = True
        elif t["aggregation"] == "count_nonzero" and prog != "universal_credit":
            if prog in caseloads.get(yr, {}):
                t["value"] = round(caseloads[yr][prog], 0)
            elif yr > LATEST_REAL_YEAR:
                t["holdout"] = True
    return targets


# ── CLI ──────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--years", nargs="+", type=int, default=TARGET_YEARS + FORECAST_YEARS
    )
    args = parser.parse_args()

    print(f"Building calibration targets for {len(args.years)} years...")
    targets = build_targets(args.years)

    from rich.console import Console
    from rich.table import Table as RichTable

    console = Console()

    by_year: dict[int, int] = {}
    for t in targets:
        by_year[t["year"]] = by_year.get(t["year"], 0) + 1

    rt = RichTable(title="Calibration targets", show_header=True)
    rt.add_column("Year", style="bold")
    rt.add_column("Targets", justify="right")
    for yr in sorted(by_year):
        rt.add_row(str(yr), str(by_year[yr]))
    rt.add_row("Total", str(len(targets)), style="bold")
    console.print(rt)

    out = {"targets": targets}
    OUT_PATH.write_text(json.dumps(out, indent=2))
    console.print(
        f"[green]Wrote {len(targets)} targets to {OUT_PATH.relative_to(REPO_ROOT)}[/green]"
    )


if __name__ == "__main__":
    main()
